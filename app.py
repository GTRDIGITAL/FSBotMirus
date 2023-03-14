# import shutil
# from time import sleep
# from random import choice
# import ctypes
# import PILasda
# import ssl
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
from openpyxl.cell import cell
from openpyxl.chart import LineChart, Reference
# import xml.etree.ElementTree as ET
from werkzeug.utils import secure_filename
from flask import flash
from flask import Flask, render_template, request, send_from_directory, after_this_request, redirect, url_for
from openpyxl.descriptors import ( 
	String,
	Sequence,
	Integer,
	)
from openpyxl.descriptors.serialisable import Serialisable
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.styles import Alignment, alignment
# from string import ascii_uppercase
import openpyxl
import datetime
# from datetime import datetime
import os
from string import ascii_uppercase
app=Flask(__name__)
app.secret_key = "GT ROMANIA Delivery Center"
var_list=[]
var_rute=[]
# app.app_context().push()
	
# @app.route("/")
# def home():
#     # text=request.form.get('client')
#     # print(text)
#     return render_template("index.html")


@app.route('/')
def FS():
	return render_template('FS.html')

@app.route('/Instructions', methods=['GET'])
def downloadPMG():
	# filepath = "D:\Projects\8. Python web apps\Test web flask\Instructions"
	return send_from_directory("/home/fsbot/storage","Instructions - FS.docx", as_attachment=True)
@app.route('/', methods=['POST', 'GET'])
def FS_process():
	company = request.form['company']
	address = request.form['address']
	vatTaxCode = request.form['code']
	regNr = request.form['registration']
	typeOfCompany = request.form['type']
	mainActivity = request.form['activity']
	year= request.form['year']
	dropdownlimba = request.form.get('limba')
	dropdownfroma = request.form.get('forma')

	if str(dropdownlimba)=="Romana(RO)":
		option=1
	else:
		option=0

	if str(dropdownfroma)=="Romana(RO)":
		option2=1
	else:
		option2=0


	folderpath="/home/fsbot/storage"
	# folderpath="D:\\FSFinal\\FSBotMirus"

	if request.method == 'POST':

		workingsblue2= Font(bold=True, italic=True, name='Tahoma', size=8,color='FFFFFF')
		lbluefill = PatternFill(start_color='7030A0',
							end_color='7030A0',
							fill_type='solid')
		grifill=PatternFill(start_color='c4d79b',end_color='c4d79b',fill_type='solid')
		yellow=PatternFill(start_color='ffff00',end_color='ffff00',fill_type='solid')
		blueFill = PatternFill(start_color='00AEAC',
							end_color='00AEAC',
							fill_type='solid')
		doubleborder = Border(bottom=Side(style='double'))
		solidborder = Border(bottom=Side(style='thick'))
		solidborderstanga = Border(left=Side(style='thin'))
		rightborder = Border(right=Side(style='thin'))
		rightdouble = Border (right=Side(style='thin'), bottom=Side(style='double'))
		rightmedium = Border (right=Side(style='thin'), bottom=Side(style='medium'))
		solidborderdreapta = Border(right=Side(style='thin'))
		solidbordersus = Border(top=Side(style='thin'))
		fontitalic = Font(name='Tahoma', size=8, bold=True, italic=True)
		font1 = Font(name='Tahoma', size=8)
		font2 = Font(name='Tahoma', size=8, bold=True)
		fontRed = Font(name='Tahoma', size=8, bold=True, color= 'FF0000')
		fontRedDiff=Font(name="Tahoma", color='FF0000', size=11, )
		fontGT = Font (name='GT Logo', size=8)
		workingsblue = Font(color='2F75B5', bold=True, name='Tahoma', size=8 )
		headers= Font(bold=True, italic=True, name='Tahoma', size=8,color='FFFFFF') 
		headersblue = PatternFill(start_color='7030A0',
						end_color='7030A0',
						fill_type='solid')
		headerspurple= PatternFill(start_color='65CDCC',
							end_color='65CDCC',
							fill_type='solid')
		total=PatternFill(start_color='DDD9C4',
						end_color='DDD9C4',
						fill_type='solid')
		greenbolditalic= Font(bold=True, italic=True,  color='C0504D', name='Tahoma', size=8)
		greenbolditalic= Font(bold=True, italic=True,  color='00af50')
		fontalb = Font(italic=True, color="bfbfbf", size=8, name='Tahoma')
		
		triald=request.files["TB"]
			

		# PBC_CY=mapping.create_sheet("Trial Balance2")
		# PBC_CY.sheet_view.showGridLines = False
		if(option2==0):
			if(option==0):
				mapping=openpyxl.load_workbook('/home/fsbot/exceltemp/SF Entitati mici_EN.xlsx')
				# mapping=openpyxl.load_workbook('C:\\Users\\denis.david\\Training materials\\SF Entitati mici_EN.xlsx')			

			else:
				mapping=openpyxl.load_workbook('/home/fsbot/exceltemp/SF Entitati mici_RO.xlsx')
				# mapping=openpyxl.load_workbook('C:\\Users\\denis.david\\Training materials\\SF Entitati mici_RO.xlsx')
			ws=mapping.active		
			TBCY = openpyxl.load_workbook(triald,data_only=True)
			TBCY1 = TBCY.active
			# PBC_CY=mapping.create_sheet("TB_PBC")
			test=mapping["Trial Balance"]
			test2=mapping["Check if manual ADJE"]


			for row in TBCY1.iter_rows():
					for cell in row:
						if cell.value=="Account":
							tbCyAcount=cell.column
							tbrow=cell.row

			for row in TBCY1.iter_rows():
				for cell in row:
					if cell.value=="Description":
						tbCyDescription=cell.column

			for row in TBCY1.iter_rows():
				for cell in row:
					if cell.value=="OB":
						tbCyOB=cell.column

			for row in TBCY1.iter_rows():

				for cell in row:
					if cell.value=="DM":
						tbCyDM=cell.column
					
			for row in TBCY1.iter_rows():
				for cell in row:
					if cell.value=="CM":
						tbCyCM=cell.column

			for row in TBCY1.iter_rows():
				for cell in row:
					if cell.value=="CB":
						tbCyCB=cell.column



			try:
				luntb=len(TBCY1[tbCyAcount])
			except:
				flash("Please insert the correct header for Account in Trial Balance file")
				return render_template("index.html")
				# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Account'")
				# sys.exit()
			try:
				Account=[b.value for b in TBCY1[tbCyAcount][tbrow:luntb+1]]
			except:
				flash("Please insert the correct header for Account in Trial Balance file")
				return render_template("index.html")
				# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Account'")
				# sys.exit()

			try:
				Description=[b.value for b in TBCY1[tbCyDescription][tbrow:luntb+1]]
			except:
				flash("Please insert the correct header for Description in Trial Balance file")
				return render_template("index.html")
				# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Description'")
				# sys.exit()
			try:
				OB=[b.value for b in TBCY1[tbCyOB][tbrow:luntb+1]]
			except:
				flash("Please insert the correct header for OB")
				return render_template("index.html")
				# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Sold Initial Debit'")
				# sys.exit()
			try:
				DM=[b.value for b in TBCY1[tbCyDM][tbrow:luntb+1]]
			except:
				flash("Please insert the correct header for DM")
				return render_template("index.html")
				# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Sold Initial Credit'")
				# sys.exit()
			try:
				CM=[b.value for b in TBCY1[tbCyCM][tbrow:luntb+1]]
			except:
				flash("Please insert the correct header for CM")
				return render_template("index.html")
				# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Rulaj Curent Debit'")
				# sys.exit()
			try:
				CB=[b.value for b in TBCY1[tbCyCB][tbrow:luntb+1]]
			except:
				flash("Please insert the correct header for CB")
				return render_template("index.html")

			for i in range(1, len(Account)+1):
				test.cell(row=i+14, column=6).value=Account[i-1]

			for i in range (1, len(Description)+1):
				test.cell(row=i+14, column=7).value= Description[i-1]

			for i in range (1, len(OB)+1):
				test.cell(row=i+14, column=8).value=OB[i-1]

			for i in range (1, len(DM)+1):
				test.cell (row=i+14, column =9).value=DM[i-1]

			for i in range (1,len(CM)+1):
				test.cell (row=i+14, column=10).value=CM[i-1]

			for i in range (1,len(CB)+1):
				test.cell (row=i+14, column=11).value=CB[i-1]


			for i in range(1, len(Account)+1):
				test.cell(row=i+14,column=2).value='=_xlfn.NUMBERVALUE(LEFT(F{0},1))'.format(i+14)	
			for i in range(1, len(Account)+1):
				test.cell(row=i+14,column=1).value='=IF(B'+str(14+i)+'<6,"BS",IF(B'+str(14+i)+'=6,"Exp","Rev"))'
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=3).value='=Left(F'+str(14+i)+',2)'
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=4).value='=Left(F'+str(14+i)+',3)'
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=5).value='=IF(F'+str(14+i)+'="121",Left(F'+str(14+i)+',3)&"0",Left(F'+str(14+i)+',4))'
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=12).value='=K'+str(14+i)+'-H'+str(14+i)+''
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=13).value='=IFERROR(L'+str(14+i)+'/H'+str(14+i)+'," ")'
			for i in range(1,len(Account)+1):
				# test.cell(row=i+14,column=14).value='''=_xlfn.IF(A'''+str(14+i)+'''="BS"'''+''',IFERROR(VLOOKUP(TRIM($E'''+str(14+i)+'),'+"'BS Mapping std'"+'!$A:$D,4,0),VLOOKUP(TRIM($D'+str(14+i)+'),'+"'BS Mapping std'"+'!$A:$D,4,0)),IFNA(VLOOKUP(TRIM($E'+str(14+i)+'),'+"'PL mapping Std'"+'!$A:$D,4,0),VLOOKUP(TRIM($D'+str(14+i)+'),'+"'PL mapping Std'"+'!$A:$D,4,0)))'
				test.cell(row=i+14,column=14).value='''=IF(A{0}="BS",IFERROR(VLOOKUP(TRIM($E{0}),'BS Mapping std'!$A:$C,3,0),VLOOKUP(TRIM($D{0}),'BS Mapping std'!$A:$C,3,0)),IFERROR(VLOOKUP(TRIM($E{0}),'PL mapping Std'!$A:$C,3,0),VLOOKUP(TRIM($D{0}),'PL mapping Std'!$A:$C,3,0)))'''.format(i+14)
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=15).value="=_xlfn.IFERROR(VLOOKUP(E"+str(14+i)+",'F30 mapping'!A:C,3,0),VLOOKUP(D"+str(14+i)+",'F30 mapping'!A:C,3,0))"
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=16).value="=_xlfn.IFERROR(IFERROR(VLOOKUP(E"+str(14+i)+",'F40 mapping'!A:C,3,0),VLOOKUP(D"+str(14+i)+",'F40 mapping'!A:C,3,0)),0)"
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=17).value="=_xlfn.IFERROR(IFERROR(VLOOKUP(E"+str(14+i)+",'F40 mapping'!A:D,4,0),VLOOKUP(D"+str(14+i)+",'F40 mapping'!A:D,4,0)),0)"
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=18).value="=_xlfn.IFERROR(IFERROR(VLOOKUP(E"+str(14+i)+",'F40 mapping'!A:E,5,0),VLOOKUP(D"+str(14+i)+",'F40 mapping'!A:E,5,0)),0)"
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=19).value="=_xlfn.IF(B"+str(14+i)+"<6,IFERROR(VLOOKUP(E"+str(14+i)+",'BS Mapping std'!A:F,6,0),VLOOKUP(D"+str(14+i)+",'BS Mapping std'!A:F,6,0)),IFERROR(VLOOKUP(E"+str(14+i)+",'PL mapping Std'!A:D,4,0),VLOOKUP(D"+str(14+i)+",'PL mapping Std'!A:D,4,0)))"
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=20).value="=_xlfn.IF(B"+str(14+i)+"<6,IFERROR(VLOOKUP(E"+str(14+i)+",'BS Mapping std'!A:G,7,0),VLOOKUP(D"+str(14+i)+",'BS Mapping std'!A:G,7,0)),IFERROR(VLOOKUP(E"+str(14+i)+",'PL mapping Std'!A:E,5,0),VLOOKUP(D"+str(14+i)+",'PL mapping Std'!A:E,5,0)))"
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=22).value='''=IF(IF(A{0}="BS",IFERROR(VLOOKUP(TRIM($E{0}),'BS Mapping std'!$A:$H,8,0),VLOOKUP(TRIM($D{0}),'BS Mapping std'!$A:$H,8,0)),IFERROR(VLOOKUP(TRIM($E{0}),'PL mapping Std'!$A:$H,8,0),VLOOKUP(TRIM($D{0}),'PL mapping Std'!$A:$H,8,0)))=0,"",IF(A{0}="BS",IFERROR(VLOOKUP(TRIM($E{0}),'BS Mapping std'!$A:$H,8,0),VLOOKUP(TRIM($D{0}),'BS Mapping std'!$A:$H,8,0)),IFERROR(VLOOKUP(TRIM($E{0}),'PL mapping Std'!$A:$H,8,0),VLOOKUP(TRIM($D{0}),'PL mapping Std'!$A:$H,8,0))))'''.format(i+14)
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=23).value="=_xlfn.IFERROR(VLOOKUP(E"+str(14+i)+",'F30 mapping'!A:D,4,0),VLOOKUP(D"+str(14+i)+",'F30 mapping'!A:D,4,0))"

			# for i in range(len(Account)+1,800):
			# 	test.cell(row=i+14,column=14).value=""
			# 	test.cell(row=i+14,column=15).value=""
			# 	test.cell(row=i+14,column=16).value=""
			# 	test.cell(row=i+14,column=17).value=""
			# 	test.cell(row=i+14,column=18).value=""
			# 	test.cell(row=i+14,column=19).value=""
			# 	test.cell(row=i+14,column=20).value=""
			# 	test.cell(row=i+14,column=21).value=""
			# 	test.cell(row=i+14,column=22).value=""
			# 	test.cell(row=i+14,column=23).value=""



			test.cell(row=1,column=2).value=company
			test.cell(row=2,column=2).value=address
			test.cell(row=3,column=2).value=vatTaxCode
			test.cell(row=4,column=2).value=regNr
			test.cell(row=5,column=2).value=typeOfCompany
			test.cell(row=6,column=2).value=mainActivity
			test.cell(row=7,column=2).value=year

			listaMapare=['161','1614','1615','1617','1618','1621','1622','1624','1625','1627','1623','1626','1661','1663','2671','2672','2673','2674','2675','2676','2677','2678','2679','308','348E','348','368','378','388','4428','445','4451','4452','4458','446','4482','4481','4511','4518','4531','4538','456','456','471','472','473','4751','4752','4753','4754','4758','481','482','581']
			listaBalanta=list(set(Account))
			listasetmapare=list(set(listaMapare))
			print(listaBalanta)
			print(listaMapare)
			val=[]
			for i in range(0,len(listaBalanta)):
				for j in range(0,len(listasetmapare)):
					if(str(listaBalanta[i])[0:3]==str(listasetmapare[j]) or str(listaBalanta[i])[0:4]==str(listasetmapare[j])):
						val.append(listaBalanta[i])

			for j  in range(0,len(val)):
				test2.cell(row=14+j,column=6).value=val[j]
			for x in range(0,len(val)):
				test2.cell(row=14+x,column=5).value='=Left(F'+str(14+x)+',4)'
				test2.cell(row=14+x,column=4).value='=Left(F'+str(14+x)+',3)'
				test2.cell(row=14+x,column=3).value='=Left(F'+str(14+x)+',2)'
				test2.cell(row=14+x,column=2).value='=Left(F'+str(14+x)+',1)'
				test2.cell(row=14+x,column=1).value='BS'
				test2.cell(row=14+x,column=7).value="=VLOOKUP(F{0},'Trial Balance'!F:G,2,0)".format(x+14)
				test2.cell(row=14+x,column=8).value="=VLOOKUP(F{0},'Trial Balance'!F:H,3,0)".format(x+14)
				test2.cell(row=14+x,column=9).value="=VLOOKUP(F{0},'Trial Balance'!F:I,4,0)".format(x+14)
				test2.cell(row=14+x,column=10).value="=VLOOKUP(F{0},'Trial Balance'!F:J,5,0)".format(x+14)
				test2.cell(row=14+x,column=11).value="=VLOOKUP(F{0},'Trial Balance'!F:K,6,0)".format(x+14)
				test2.cell(row=x+14,column=12).value='=K'+str(14+x)+'-H'+str(14+x)+''
				test2.cell(row=x+14,column=13).value='=IFERROR(L'+str(14+x)+'/H'+str(14+x)+'," ")'
				test2.cell(row=x+14,column=14).value='''=IF(A{0}="BS",IFERROR(VLOOKUP(TRIM($E{0}),'BS Mapping std'!$A:$D,4,0),VLOOKUP(TRIM($D{0}),'BS Mapping std'!$A:$D,4,0)),IFERROR(VLOOKUP(TRIM($E{0}),'PL mapping Std'!$A:$D,4,0),VLOOKUP(TRIM($D{0}),'PL mapping Std'!$A:$D,4,0)))'''.format(x+14)
				# test2.cell(row=14+x,column=14).value='Bifunctional - Please asses'
					# else:
					# 	print("nu sunt la fel")


			print(len(listaBalanta))
			print(len(listaMapare))

			f10=mapping["1. F10"]
			f20=mapping["2. F20"]
			f30=mapping["3. F30"]
			f40=mapping["4. F40"]
			# soce=mapping["5. SOCE"]
			# socf=mapping["6.SOCF"]
			n3nca=mapping["N3 - NCA"]
			n4inv=mapping["N4 - Inventories"]
			n7cash=mapping["N7 - Cash"]
			n5tr=mapping["N5 - TR"]
			n9tp=mapping["N9 - TP"]
			n10prov=mapping["N10 - Provisions"]
			n15pers=mapping["N15 - Personnel"]
			n16opex=mapping["N16 - Other OPEX"]

			f10.print_area="A11:E132"
			f20.print_area="A11:E91"
			f30.print_area="A17:F292"
			f40.print_area="A10:H79"
			# soce.print_area="E8:L38"
			# socf.print_area="A7:C51"
			n3nca.print_area="A10:O73"
			n4inv.print_area="A10:G24"
			n5tr.print_area="A10:F47"
			n7cash.print_area="A10:C19"
			n9tp.print_area="A10:G41"
			n10prov.print_area="A10:G22"
			n15pers.print_area="A10:C27"
			n16opex.print_area="A10:C30"
			mapping.save(str(folderpath)+"/Financial Statements-"+str(company)+".xlsx")
			# return send_from_directory(folderpath,"Financial Statements-"+str(company)+".xlsx",as_attachment=True)

		else:
			if(option==0):
				mapping=openpyxl.load_workbook('/home/fsbot/exceltemp/Template FS ENG.xlsx')
				# mapping=openpyxl.load_workbook('C:\\Users\\denis.david\\Training materials\\Template FS ENG.xlsx')			

			else:
				mapping=openpyxl.load_workbook('/home/fsbot/exceltemp/Template FS RO.xlsx')
				# mapping=openpyxl.load_workbook('C:\\Users\\denis.david\\Training materials\\Template FS RO.xlsx')
			ws=mapping.active		
			TBCY = openpyxl.load_workbook(triald)
			TBCY1 = TBCY.active
			# PBC_CY=mapping.create_sheet("TB_PBC")
			test=mapping["Trial Balance"]
			test2=mapping["Check if manual ADJE"]


			for row in TBCY1.iter_rows():
					for cell in row:
						if cell.value=="Account":
							tbCyAcount=cell.column
							tbrow=cell.row

			for row in TBCY1.iter_rows():
				for cell in row:
					if cell.value=="Description":
						tbCyDescription=cell.column

			for row in TBCY1.iter_rows():
				for cell in row:
					if cell.value=="OB":
						tbCyOB=cell.column

			for row in TBCY1.iter_rows():

				for cell in row:
					if cell.value=="DM":
						tbCyDM=cell.column
					
			for row in TBCY1.iter_rows():
				for cell in row:
					if cell.value=="CM":
						tbCyCM=cell.column

			for row in TBCY1.iter_rows():
				for cell in row:
					if cell.value=="CB":
						tbCyCB=cell.column



			try:
				luntb=len(TBCY1[tbCyAcount])
			except:
				flash("Please insert the correct header for Account in Trial Balance file")
				return render_template("index.html")
				# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Account'")
				# sys.exit()
			try:
				Account=[b.value for b in TBCY1[tbCyAcount][tbrow:luntb+1]]
			except:
				flash("Please insert the correct header for Account in Trial Balance file")
				return render_template("index.html")
				# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Account'")
				# sys.exit()

			try:
				Description=[b.value for b in TBCY1[tbCyDescription][tbrow:luntb+1]]
			except:
				flash("Please insert the correct header for Description in Trial Balance file")
				return render_template("index.html")
				# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Description'")
				# sys.exit()
			try:
				OB=[b.value for b in TBCY1[tbCyOB][tbrow:luntb+1]]
			except:
				flash("Please insert the correct header for OB")
				return render_template("index.html")
				# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Sold Initial Debit'")
				# sys.exit()
			try:
				DM=[b.value for b in TBCY1[tbCyDM][tbrow:luntb+1]]
			except:
				flash("Please insert the correct header for DM")
				return render_template("index.html")
				# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Sold Initial Credit'")
				# sys.exit()
			try:
				CM=[b.value for b in TBCY1[tbCyCM][tbrow:luntb+1]]
			except:
				flash("Please insert the correct header for CM")
				return render_template("index.html")
				# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Rulaj Curent Debit'")
				# sys.exit()
			try:
				CB=[b.value for b in TBCY1[tbCyCB][tbrow:luntb+1]]
			except:
				flash("Please insert the correct header for CB")
				return render_template("index.html")

			for i in range(1, len(Account)+1):
				test.cell(row=i+14, column=6).value=str(Account[i-1])

			for i in range (1, len(Description)+1):
				test.cell(row=i+14, column=7).value= Description[i-1]

			for i in range (1, len(OB)+1):
				test.cell(row=i+14, column=8).value=OB[i-1]

			for i in range (1, len(DM)+1):
				test.cell (row=i+14, column =9).value=DM[i-1]

			for i in range (1,len(CM)+1):
				test.cell (row=i+14, column=10).value=CM[i-1]

			for i in range (1,len(CB)+1):
				test.cell (row=i+14, column=11).value=CB[i-1]


			for i in range(1, len(Account)+1):
				test.cell(row=i+14,column=2).value='=_xlfn.NUMBERVALUE(LEFT(F{0},1))'.format(i+14)	
			for i in range(1, len(Account)+1):
				test.cell(row=i+14,column=1).value='=IF(B'+str(14+i)+'<6,"BS",IF(B'+str(14+i)+'=6,"Exp","Rev"))'
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=3).value='=Left(F'+str(14+i)+',2)'
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=4).value='=Left(F'+str(14+i)+',3)'
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=5).value='=IF(F'+str(14+i)+'="121",Left(F'+str(14+i)+',3)&"0",Left(F'+str(14+i)+',4))'
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=12).value='=K'+str(14+i)+'-H'+str(14+i)+''
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=13).value='=IFERROR(L'+str(14+i)+'/H'+str(14+i)+'," ")'
			for i in range(1,len(Account)+1):
				# test.cell(row=i+14,column=14).value='''=_xlfn.IF(A'''+str(14+i)+'''="BS"'''+''',IFERROR(VLOOKUP(TRIM($E'''+str(14+i)+'),'+"'BS Mapping std'"+'!$A:$D,4,0),VLOOKUP(TRIM($D'+str(14+i)+'),'+"'BS Mapping std'"+'!$A:$D,4,0)),IFNA(VLOOKUP(TRIM($E'+str(14+i)+'),'+"'PL mapping Std'"+'!$A:$D,4,0),VLOOKUP(TRIM($D'+str(14+i)+'),'+"'PL mapping Std'"+'!$A:$D,4,0)))'
				test.cell(row=i+14,column=14).value='''=IF(A{0}="BS",IFERROR(VLOOKUP(TRIM($E{0}),'BS Mapping std'!$A:$D,4,0),VLOOKUP(TRIM($D{0}),'BS Mapping std'!$A:$D,4,0)),IFERROR(VLOOKUP(TRIM($E{0}),'PL mapping Std'!$A:$D,4,0),VLOOKUP(TRIM($D{0}),'PL mapping Std'!$A:$D,4,0)))'''.format(i+14)
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=15).value="=_xlfn.IFERROR(VLOOKUP(E"+str(14+i)+",'F30 mapping'!A:C,3,0),VLOOKUP(D"+str(14+i)+",'F30 mapping'!A:C,3,0))"
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=16).value="=_xlfn.IFERROR(IFERROR(VLOOKUP(E"+str(14+i)+",'F40 mapping'!A:C,3,0),VLOOKUP(D"+str(14+i)+",'F40 mapping'!A:C,3,0)),0)"
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=17).value="=_xlfn.IFERROR(IFERROR(VLOOKUP(E"+str(14+i)+",'F40 mapping'!A:D,4,0),VLOOKUP(D"+str(14+i)+",'F40 mapping'!A:D,4,0)),0)"
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=18).value="=_xlfn.IFERROR(IFERROR(VLOOKUP(E"+str(14+i)+",'F40 mapping'!A:E,5,0),VLOOKUP(D"+str(14+i)+",'F40 mapping'!A:E,5,0)),0)"
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=19).value="=_xlfn.IF(B"+str(14+i)+"<6,IFERROR(VLOOKUP(E"+str(14+i)+",'BS Mapping std'!A:E,5,0),VLOOKUP(D"+str(14+i)+",'BS Mapping std'!A:E,5,0)),IFERROR(VLOOKUP(E"+str(14+i)+",'PL mapping Std'!A:F,6,0),VLOOKUP(D"+str(14+i)+",'PL mapping Std'!A:F,6,0)))"
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=20).value="=_xlfn.IF(B"+str(14+i)+"<6,IFERROR(VLOOKUP(E"+str(14+i)+",'BS Mapping std'!A:F,6,0),VLOOKUP(D"+str(14+i)+",'BS Mapping std'!A:F,6,0)),IFERROR(VLOOKUP(E"+str(14+i)+",'PL mapping Std'!A:G,7,0),VLOOKUP(D"+str(14+i)+",'PL mapping Std'!A:G,7,0)))"
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=22).value='''=IF(IF(A{0}="BS",IFERROR(VLOOKUP(TRIM($E{0}),'BS Mapping std'!$A:$H,8,0),VLOOKUP(TRIM($D{0}),'BS Mapping std'!$A:$H,8,0)),IFERROR(VLOOKUP(TRIM($E{0}),'PL mapping Std'!$A:$H,8,0),VLOOKUP(TRIM($D{0}),'PL mapping Std'!$A:$H,8,0)))=0,"",IF(A{0}="BS",IFERROR(VLOOKUP(TRIM($E{0}),'BS Mapping std'!$A:$H,8,0),VLOOKUP(TRIM($D{0}),'BS Mapping std'!$A:$H,8,0)),IFERROR(VLOOKUP(TRIM($E{0}),'PL mapping Std'!$A:$H,8,0),VLOOKUP(TRIM($D{0}),'PL mapping Std'!$A:$H,8,0))))'''.format(i+14)
			for i in range(1,len(Account)+1):
				test.cell(row=i+14,column=23).value="=_xlfn.IFERROR(VLOOKUP(E"+str(14+i)+",'F30 mapping'!A:D,4,0),VLOOKUP(D"+str(14+i)+",'F30 mapping'!A:D,4,0))"

			# for i in range(len(Account)+1,800):
			# 	test.cell(row=i+14,column=14).value=""
			# 	test.cell(row=i+14,column=15).value=""
			# 	test.cell(row=i+14,column=16).value=""
			# 	test.cell(row=i+14,column=17).value=""
			# 	test.cell(row=i+14,column=18).value=""
			# 	test.cell(row=i+14,column=19).value=""
			# 	test.cell(row=i+14,column=20).value=""
			# 	test.cell(row=i+14,column=21).value=""
			# 	test.cell(row=i+14,column=22).value=""
			# 	test.cell(row=i+14,column=23).value=""



			test.cell(row=1,column=2).value=company
			test.cell(row=2,column=2).value=address
			test.cell(row=3,column=2).value=vatTaxCode
			test.cell(row=4,column=2).value=regNr
			test.cell(row=5,column=2).value=typeOfCompany
			test.cell(row=6,column=2).value=mainActivity
			test.cell(row=7,column=2).value=year

			listaMapare=['161','1614','1615','1617','1618','1621','1622','1624','1625','1627','1623','1626','1661','1663','2671','2672','2673','2674','2675','2676','2677','2678','2679','308','348E','348','368','378','388','4428','445','4451','4452','4458','446','4482','4481','4511','4518','4531','4538','456','456','471','472','473','4751','4752','4753','4754','4758','481','482','581']
			listaBalanta=list(set(Account))
			listasetmapare=list(set(listaMapare))
			print(listaBalanta)
			print(listaMapare)
			val=[]
			for i in range(0,len(listaBalanta)):
				for j in range(0,len(listasetmapare)):
					if(str(listaBalanta[i])[0:3]==str(listasetmapare[j]) or str(listaBalanta[i])[0:4]==str(listasetmapare[j])):
						val.append(listaBalanta[i])

			for j  in range(0,len(val)):
				test2.cell(row=14+j,column=6).value=val[j]
			for x in range(0,len(val)):
				test2.cell(row=14+x,column=5).value='=Left(F'+str(14+x)+',4)'
				test2.cell(row=14+x,column=4).value='=Left(F'+str(14+x)+',3)'
				test2.cell(row=14+x,column=3).value='=Left(F'+str(14+x)+',2)'
				test2.cell(row=14+x,column=2).value='=Left(F'+str(14+x)+',1)'
				test2.cell(row=14+x,column=1).value='BS'
				test2.cell(row=14+x,column=7).value="=VLOOKUP(F{0},'Trial Balance'!F:G,2,0)".format(x+14)
				test2.cell(row=14+x,column=8).value="=VLOOKUP(F{0},'Trial Balance'!F:H,3,0)".format(x+14)
				test2.cell(row=14+x,column=9).value="=VLOOKUP(F{0},'Trial Balance'!F:I,4,0)".format(x+14)
				test2.cell(row=14+x,column=10).value="=VLOOKUP(F{0},'Trial Balance'!F:J,5,0)".format(x+14)
				test2.cell(row=14+x,column=11).value="=VLOOKUP(F{0},'Trial Balance'!F:K,6,0)".format(x+14)
				test2.cell(row=x+14,column=12).value='=K'+str(14+x)+'-H'+str(14+x)+''
				test2.cell(row=x+14,column=13).value='=IFERROR(L'+str(14+x)+'/H'+str(14+x)+'," ")'
				test2.cell(row=x+14,column=14).value="=VLOOKUP(F{0},'Trial Balance'!F:N,9,0)".format(x+14)
				test2.cell(row=14+x,column=15).value='Bifunctional - Please asses'
					# else:
					# 	print("nu sunt la fel")


			print(len(listaBalanta))
			print(len(listaMapare))

			f10=mapping["1. F10"]
			f20=mapping["2. F20"]
			f30=mapping["3. F30"]
			f40=mapping["4. F40"]
			soce=mapping["5. SOCE"]
			socf=mapping["6.SOCF"]
			n3nca=mapping["N3 - NCA"]
			n4inv=mapping["N4 - Inventories"]
			n5tr=mapping["N5 - TR"]
			n7cash=mapping["N7 - Cash"]
			n9tp=mapping["N9 - TP"]
			n10prov=mapping["N10 - Provisions"]
			n15pers=mapping["N15 - Personnel"]
			n16opex=mapping["N16 - Other OPEX"]

			f10.print_area="A11:E132"
			f20.print_area="A11:E91"
			f30.print_area="A17:F292"
			f40.print_area="A10:H79"
			soce.print_area="E8:L38"
			socf.print_area="A7:C51"
			n3nca.print_area="A10:O73"
			n4inv.print_area="A10:G24"
			n5tr.print_area="A10:F47"
			n7cash.print_area="A10:C19"
			n9tp.print_area="A10:G41"
			n10prov.print_area="A10:G22"
			n15pers.print_area="A10:C27"
			n16opex.print_area="A10:C30"
			mapping.save(str(folderpath)+"/Financial Statements-"+str(company)+".xlsx")
	return send_from_directory(folderpath,"Financial Statements-"+str(company)+".xlsx",as_attachment=True)


@app.route('/TrialBalances/Instructions', methods=['GET'])
def downloadTB():
		filepath = "/home/fsbot/storage"
 
		return send_from_directory(filepath,"Instructions - Trial Balance.docx", as_attachment=True)
@app.route('/TrialBalances/GT3SjGyxpbcxV35PeSUpKJQIOgY')
def TB():
	return render_template('TB.html')
@app.route('/TrialBalances/GT3SjGyxpbcxV35PeSUpKJQIOgY', methods=['POST', 'GET'])
def TB_process():

	namec = request.form['client']
	ant= datetime.strptime(
					 request.form['yearEnd'],
					 '%Y-%m-%d')
	threshol = request.form['threshold']
	preparedBy1 = request.form['preparedBy']
	isChecked1=request.form.get("Stdmapp")
	isChecked2=request.form.get("forml")
	isChecked3=request.form.get("forms")
	isChecked4=request.form.get("pyEx")
	# denis=datetime.datetime.now()


		# if isChecked1=="": #daca e bifat
	#     isChecked=1
	# else:
	#     isChecked=0
	
	# folderpath="/home/auditappnexia/output/tb"
	folderpath="/home/fsbot/storage"

	def make_archive(source, destination):
		base = os.path.basename(destination)
		name = base.split('.')[0]
		format = base.split('.')[1]
		archive_from = os.path.dirname(source)
		archive_to = os.path.basename(source.strip(os.sep))
		shutil.make_archive(name, format, archive_from, archive_to)
		shutil.move('%s.%s'%(name,format), destination)
	# yearEnd = str(request.form['yearEnd'])
	# processed_text = client.upper()
	# fisier=request.files.get('monthlyTB')
	if request.method == 'POST':
		workingsblue2= Font(bold=True, italic=True, name='Tahoma', size=8,color='FFFFFF')
		lbluefill = PatternFill(start_color='7030A0',
							end_color='7030A0',
							fill_type='solid')
		grifill=PatternFill(start_color='c4d79b',end_color='c4d79b',fill_type='solid')
		yellow=PatternFill(start_color='ffff00',end_color='ffff00',fill_type='solid')
		blueFill = PatternFill(start_color='00AEAC',
							end_color='00AEAC',
							fill_type='solid')
		doubleborder = Border(bottom=Side(style='double'))
		solidborder = Border(bottom=Side(style='thick'))
		solidborderstanga = Border(left=Side(style='thin'))
		rightborder = Border(right=Side(style='thin'))
		rightdouble = Border (right=Side(style='thin'), bottom=Side(style='double'))
		rightmedium = Border (right=Side(style='thin'), bottom=Side(style='medium'))
		solidborderdreapta = Border(right=Side(style='thin'))
		solidbordersus = Border(top=Side(style='thin'))
		fontitalic = Font(name='Tahoma', size=8, bold=True, italic=True)
		font = Font(name='Tahoma', size=8, bold=True)
		font1 = Font(name='Tahoma', size=8)
		font2 = Font(name='Tahoma', size=10, bold=True)
		fontRed = Font(name='Tahoma', size=10, bold=True, color= 'FF0000')
		fontRedDiff=Font(name="Tahoma", color='FF0000', size=11, )
		fontGT = Font (name='GT Logo', size=8)
		workingsblue = Font(color='2F75B5', bold=True, name='Tahoma', size=8 )
		headers= Font(bold=True, italic=True, name='Tahoma', size=8,color='FFFFFF') 
		headersblue = PatternFill(start_color='7030A0',
						end_color='7030A0',
						fill_type='solid')
		headerspurple= PatternFill(start_color='65CDCC',
							end_color='65CDCC',
							fill_type='solid')
		total=PatternFill(start_color='DDD9C4',
						end_color='DDD9C4',
						fill_type='solid')
		greenbolditalic= Font(bold=True, italic=True,  color='C0504D', name='Tahoma', size=8)
		greenbolditalic= Font(bold=True, italic=True,  color='00af50')
		fontalb = Font(italic=True, color="bfbfbf", size=8, name='Tahoma')
		trialc=request.files["trialBalCYPBC"]
		trialp=request.files["trialBalPYPBC"]
		TBCY = openpyxl.load_workbook(trialc,data_only=True)
		TBCY1 = TBCY.active

		# if isChecked4=="":
		# 	try:
		
		"Open files"




		"Iterate from imported PBC's:"


		'Iterate from CY TB'

		for row in TBCY1.iter_rows():
				for cell in row:
					if cell.value=="Account":
						tbCyAcount=cell.column
						tbrow=cell.row

		for row in TBCY1.iter_rows():
			for cell in row:
				if cell.value=="Description":
					tbCyDescription=cell.column

		for row in TBCY1.iter_rows():
			for cell in row:
				if cell.value=="SID":
					tbCySID=cell.column

		for row in TBCY1.iter_rows():

			for cell in row:
				if cell.value=="SIC":
					tbCySIC=cell.column
				
		for row in TBCY1.iter_rows():
			for cell in row:
				if cell.value=="RCD":
					tbCyRCD=cell.column

		for row in TBCY1.iter_rows():
			for cell in row:
				if cell.value=="RCC":
					tbCyRCC=cell.column

		for row in TBCY1.iter_rows():
			for cell in row:
				if cell.value=="SFD":
					tbCySFD=cell.column

		for row in TBCY1.iter_rows():
			for cell in row:
				if cell.value=="SFC":
					tbCySFC=cell.column


		try:
			luntb=len(TBCY1[tbCyAcount])
		except:
			flash("Please insert the correct header for Account in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Account'")
			# sys.exit()
		try:
			Account=[b.value for b in TBCY1[tbCyAcount][tbrow:luntb+1]]
		except:
			flash("Please insert the correct header for Account in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Account'")
			# sys.exit()

		try:
			Description=[b.value for b in TBCY1[tbCyDescription][tbrow:luntb+1]]
		except:
			flash("Please insert the correct header for Description in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Description'")
			# sys.exit()
		try:
			SID=[b.value for b in TBCY1[tbCySID][tbrow:luntb+1]]
		except:
			flash("Please insert the correct header for Sold Initial Debit in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Sold Initial Debit'")
			# sys.exit()
		try:
			SIC=[b.value for b in TBCY1[tbCySIC][tbrow:luntb+1]]
		except:
			flash("Please insert the correct header for Sold Initial Credit in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Sold Initial Credit'")
			# sys.exit()
		try:
			RCD=[b.value for b in TBCY1[tbCyRCD][tbrow:luntb+1]]
		except:
			flash("Please insert the correct header for Rulaj Curent Debit in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Rulaj Curent Debit'")
			# sys.exit()
		try:
			RCC=[b.value for b in TBCY1[tbCyRCC][tbrow:luntb+1]]
		except:
			flash("Please insert the correct header for Rulaj Curent Credit in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Rulaj Curent credit'")
			# sys.exit()
		try:
			SFD=[b.value for b in TBCY1[tbCySFD][tbrow:luntb+1]]
		except:
			flash("Please insert the correct header for Sold Final Debit in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Sold Final Debit'")
			# sys.exit()
		try: 
			SFC=[b.value for b in TBCY1[tbCySFC][tbrow:luntb+1]]
		except:
			flash("Please insert the correct header for Sold Final Credit in Trial Balance file")
			return render_template("index.html")
			# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Sold Final Credit'")
			# sys.exit()
		"Create CY PBC"

		if isChecked4=="":
			TBPY = openpyxl.load_workbook(trialp,data_only=True)
			TBPY1 = TBPY.active
			try:
				for row in TBPY1.iter_rows():
						for cell in row:
							if cell.value=="Account":
								tbPyAcount=cell.column
								tbPYrow=cell.row

				for row in TBPY1.iter_rows():
					for cell in row:
						if cell.value=="Description":
							tbPyDescription=cell.column


				for row in TBPY1.iter_rows():
					for cell in row:
						if cell.value=="CB":
							tbPySFD=cell.column


				try:
					luntbp=len(TBPY1[tbPyAcount])
				except:
					flash("Please insert the correct header for Account in Trial Balance Prior Year file")
					return render_template("index.html")
					# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Account'")
					# sys.exit()
				try:
					Accountp=[b.value for b in TBPY1[tbPyAcount][tbPYrow:luntb+1]]
				except:
					flash("Please insert the correct header for Account in Trial Balance Prior Year file")
					return render_template("index.html")
					# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Account'")
					# sys.exit()

				try:
					Descriptionp=[b.value for b in TBPY1[tbPyDescription][tbPYrow:luntb+1]]
				except:
					flash("Please insert the correct header for Description in Trial Balance Prior Year file")
					return render_template("index.html")
					# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Description'")
					# sys.exit()

				try:
					CB=[b.value for b in TBPY1[tbPySFD][tbPYrow:luntb+1]]
				except:
					flash("Please insert the correct header for CB in Trial Balance Prior Year file")
					return render_template("index.html")
					# messagebox.showerror("Error", "File: Trial Balance. Please insert the correct header for 'Sold Final Debit'")
					# sys.exit()
			except:
				pass
		
		# if isChecked1=="":
		# 	mapp=request.files["Mapping"]
		output=openpyxl.Workbook()
		# else:
		# 	if isChecked2=="":
		# 		output=openpyxl.load_workbook("/home/auditappnexia/output/otherfiles/Mapping Forma Lunga.xlsx",data_only=True)
		# 	else:
		# 		output=openpyxl.load_workbook("/home/auditappnexia/output/otherfiles/Mapping Forma Scurta.xlsx",data_only=True)

		PBC_CY =output.create_sheet("PBC_CY")

		PBC_CY.cell(row=1, column=1).value="Class"
		PBC_CY.cell(row=1, column=2).value="Synt3"
		PBC_CY.cell(row=1, column=3).value="Synt4"
		PBC_CY.cell(row=1, column=4).value="Account"
		PBC_CY.cell(row=1, column=5).value="Description"
		PBC_CY.cell(row=1, column=6).value="SID"
		PBC_CY.cell(row=1, column=7).value="SIC"
		PBC_CY.cell(row=1, column=8).value="RCD"
		PBC_CY.cell(row=1, column=9).value="RCC"
		PBC_CY.cell(row=1, column=10).value="SFD"
		PBC_CY.cell(row=1, column=11).value="SFC"


		for i in range (1,10):
			PBC_CY.cell (row=1, column=i).border=doubleborder
			PBC_CY.cell (row=1, column=i).font=font2


		for i in range(1, len(Account)+1):
			PBC_CY.cell(row=i+1, column=4).value=Account[i-1]

		for i in range (1, len(Description)+1):
			PBC_CY.cell(row=i+1, column=5).value= Description[i-1]

		for i in range (1, len(SID)+1):
			PBC_CY.cell(row=i+1, column=6).value=SID[i-1]

		for i in range (1, len(SIC)+1):
			PBC_CY.cell (row=i+1, column =7).value=SIC[i-1]

		for i in range (1,len(RCD)+1):
			PBC_CY.cell (row=i+1, column=8).value=RCD[i-1]

		for i in range (1,len(RCC)+1):
			PBC_CY.cell (row=i+1, column=9).value=RCC[i-1]

		for i in range (1,len(SFD)+1):
			PBC_CY.cell (row=i+1, column=10).value=SFD[i-1]

		for i in range(1,len(SFC)+1):
			PBC_CY.cell (row=i+1, column=11).value=SFC[i-1]

		for i in range (1,12):
			PBC_CY.cell(row=1, column=i).font=font2
			PBC_CY.cell(row=1, column=i).border=doubleborder
			PBC_CY.cell(row=1, column=i).fill=blueFill

		for i in range (1, len(SFD)+1):
			for j in range (6, 12):
				PBC_CY.cell(row=i+1, column=j).number_format='#,##0_);(#,##0)'


		for i in range(1, len(Account)+1):
			PBC_CY.cell(row=i+1,column=1).value='=Left(D{0},1)'.format(i+1)
		for i in range(1,len(Account)+1):
			PBC_CY.cell(row=i+1,column=2).value='=Left(D{0},3)'.format(i+1)
		for i in range(1,len(Account)+1):
			PBC_CY.cell(row=i+1,column=3).value='=Left(D{0},4)'.format(i+1)


		PBC_PY =output.create_sheet("PBC_PY")

		if isChecked4=="":
			try:
				PBC_PY.cell(row=1, column=1).value="Class"
				PBC_PY.cell(row=1, column=2).value="Synt3"
				PBC_PY.cell(row=1, column=3).value="Synt4"
				PBC_PY.cell(row=1, column=4).value="Account"
				PBC_PY.cell(row=1, column=5).value="Description"
				PBC_PY.cell(row=1, column=6).value="CB"


				for i in range (1,8):
					PBC_PY.cell (row=1, column=i).border=doubleborder
					PBC_PY.cell (row=1, column=i).font=font2


				for i in range(1, len(Accountp)+1):
						PBC_PY.cell(row=i+1, column=4).value=Accountp[i-1]

				for i in range (1, len(Descriptionp)+1):
					PBC_PY.cell(row=i+1, column=5).value= Descriptionp[i-1]


				for i in range (1,len(CB)+1):
					PBC_PY.cell (row=i+1, column=6).value=CB[i-1]

				for i in range (1,8):
					PBC_PY.cell(row=1, column=i).font=font2
					PBC_PY.cell(row=1, column=i).border=doubleborder
					PBC_PY.cell(row=1, column=i).fill=blueFill

				for i in range (1, len(CB)+1):
					for j in range (6, 8):
						PBC_PY.cell(row=i+1, column=j).number_format='#,##0_);(#,##0)'


				for i in range(1, len(Accountp)+1):
					PBC_PY.cell(row=i+1,column=1).value='=Left(D{0},1)'.format(i+1)
				for i in range(1,len(Accountp)+1):
					PBC_PY.cell(row=i+1,column=2).value='=Left(D{0},3)'.format(i+1)
				for i in range(1,len(Accountp)+1):
					PBC_PY.cell(row=i+1,column=3).value='=Left(D{0},4)'.format(i+1)
			except:
				pass

		"Define F10 Worksheet"

		F10TB=output.create_sheet("F_10_Trial_Balance")
		F10TB.sheet_view.showGridLines = False
		F10TB.cell(row=1, column=1).value="Client:"
		F10TB.cell(row=1, column=1).font=font
		F10TB.cell(row=1, column=2).value=namec
		F10TB.cell(row=1, column=2).font=font


		F10TB.cell(row=2, column=1).value="Period end:"
		F10TB.cell(row=2, column=1).font=font
		F10TB.cell(row=2, column=2).value=ant
		F10TB.cell(row=2, column=2).font=font
		F10TB.cell(row=2, column=2).number_format="mm/dd/yyyy"


		F10TB.cell(row=1, column=11).value="Prepared by:"
		F10TB.cell(row=1, column=11).font=font
		F10TB.cell(row=1, column=12).value=preparedBy1
		F10TB.cell(row=1, column=12).font=font

		F10TB.cell(row=2, column=11).value="Date:"
		F10TB.cell(row=2, column=11).font=font
		# F10TB.cell(row=2, column=12).value=datetime.datetime.now()
		F10TB.cell(row=2, column=12).number_format="mm/dd/yyyy"
		F10TB.cell(row=2, column=12).alignment=Alignment(horizontal='left')

		F10TB.cell(row=3, column=11).value="Ref:"
		F10TB.cell(row=3, column=11).font=font
		F10TB.cell(row=3, column=12).value="F10"
		F10TB.cell(row=3, column=12).font=fontRed

		for i in range(1,4):
			F10TB.cell(row=i, column=11).alignment=Alignment(horizontal='right')

		F10TB.cell(row=4, column=2).value="Trial Balance"
		F10TB.cell(row=4, column=2).font=font

		F10TB.cell(row=6, column=1).value="Work done:"
		F10TB.cell(row=6, column=1).font=font


		F10TB.cell(row=8, column=1).value="(to be adjusted for P&L variation; e.g. if YE is different of 31.12)"
		F10TB.cell(row=8, column=1).font=Font(name='Tahoma', size=8, italic=True)




		F10TB.cell(row=14, column=1).value="Class"
		F10TB.cell(row=14, column=2).value="Synt 1"
		F10TB.cell(row=14, column=3).value="Synt 3"
		F10TB.cell(row=14, column=4).value="Synt 4"
		F10TB.cell(row=14, column=5).value="Account"
		F10TB.cell(row=14, column=6).value="Description"
		F10TB.cell(row=14, column=7).value="OB"
		F10TB.cell(row=14, column=8).value="DM"
		F10TB.cell(row=14, column=9).value="CM"
		F10TB.cell(row=14, column=10).value="CB"
		F10TB.cell(row=14, column=11).value="Check"
		F10TB.cell(row=14, column=13).value="Abs CB-OB"
		F10TB.cell(row=14, column=14).value="VAR %"

		# F10TB.cell(row=14, column=16).value="OMF Row"
		# F10TB.cell(row=14, column=17).value="OMF Description"
		# F10TB.cell(row=14, column=18).value="LS"

		F10TB.cell(row=14, column=16).value="Check OB"



		for i in range(1,12):
			F10TB.cell(row=14, column=i).font=font2
			F10TB.cell(row=14, column=i).fill=blueFill
			F10TB.cell(row=14, column=i).border=doubleborder
			F10TB.cell(row=14, column=i).alignment=Alignment(horizontal='left')

		for i in range(13,15):
			F10TB.cell(row=14, column=i).font=font2
			F10TB.cell(row=14, column=i).fill=blueFill
			F10TB.cell(row=14, column=i).border=doubleborder
			F10TB.cell(row=14, column=i).alignment=Alignment(horizontal='left')

		# F10TB.cell(row=14, column=16).font=font2
		# F10TB.cell(row=14, column=16).fill=blueFill
		# F10TB.cell(row=14, column=16).border=doubleborder
		# F10TB.cell(row=14, column=16).alignment=Alignment(horizontal='left')

		# F10TB.cell(row=14, column=17).font=font2
		# F10TB.cell(row=14, column=17).fill=blueFill
		# F10TB.cell(row=14, column=17).border=doubleborder
		# F10TB.cell(row=14, column=17).alignment=Alignment(horizontal='left')

		# F10TB.cell(row=14, column=18).font=font2
		# F10TB.cell(row=14, column=18).fill=blueFill
		# F10TB.cell(row=14, column=18).border=doubleborder
		F10TB.cell(row=14, column=18).alignment=Alignment(horizontal='left')

		F10TB.cell(row=8, column=6).value="Check BS"
		F10TB.cell(row=9, column=6).value="Revenues"
		F10TB.cell(row=10, column=6).value="Expenses"
		F10TB.cell(row=11, column=6).value="Result"
		F10TB.cell(row=11, column=6).border=doubleborder
		F10TB.cell(row=12, column=6).value="Check"

		for i in range (8,13):
			F10TB.cell(row=i, column=6).font=font
			F10TB.cell(row=i, column=6).alignment=Alignment(horizontal='right')

		F10TB.cell(row=8, column=7).value = '=SUMIF(A:A,"BS",G:G)'
		F10TB.cell(row=11, column=7).border=doubleborder

		for i in range (8,14):
			F10TB.cell(row=i, column=7).number_format='#,##0_);(#,##0)'


		F10TB.cell(row=8, column=10).value = '=SUMIF(A:A,"BS",J:J)'
		F10TB.cell(row=9, column=10).value='=SUMIF(B:B,"7",J:J)'
		F10TB.cell(row=10, column=10).value='=SUMIF(B:B,"6",J:J)'
		F10TB.cell(row=11, column=10).value='=SUMIF(C:C,"121",J:J)'
		F10TB.cell(row=11, column=10).border=doubleborder
		F10TB.cell(row=12, column=10).value="=SUM(J9:J10)-J11"
		F10TB.cell(row=12, column=10).font=fontRed

		F10TB.cell(row=8, column=7).value = '=SUMIF(A:A,"BS",G:G)'
		F10TB.cell(row=9, column=7).value='=SUMIF(B:B,"7",G:G)'
		F10TB.cell(row=10, column=7).value='=SUMIF(B:B,"6",G:G)'
		F10TB.cell(row=11, column=7).value='=SUMIF(C:C,"121",G:G)'
		F10TB.cell(row=11, column=7).border=doubleborder
		F10TB.cell(row=12, column=7).value="=SUM(G9:G10)-G11"
		F10TB.cell(row=12, column=7).font=fontRed



		F10TB.cell(row=13, column=12).value="=SUM(K:K)"
		F10TB.cell(row=13, column=12).font=fontRed

		F10TB.cell(row=13,column=11).value="Total diff:"
		F10TB.cell(row=13,column=11).alignment=Alignment(horizontal='right')

		for i in range (8,13):
			F10TB.cell(row=i, column=10).number_format='#,##0_);(#,##0)'



		for i in range (1,10):
			F10TB.cell(row=i, column=15).number_format='#,##0_);(#,##0)'

		for i in range(14,16):
			F10TB.cell(row=1, column=i).font=font2
			F10TB.cell(row=1, column=i).fill=blueFill
			F10TB.cell(row=1, column=i).border=doubleborder
			F10TB.cell(row=1, column=i).alignment=Alignment(horizontal='left')

		"Importing Data"

		if isChecked4=="":
			try:

				acc=Account+Accountp

				mylist2 = list(dict.fromkeys(acc))
				mylist=[]
				for xxx in range(0,len(mylist2)):
					mylist.append(str(mylist2[xxx]))
				mylist.sort()

				print(mylist)
			except:
				# acc=Account
				mylist2 = list(dict.fromkeys(Account))
				mylist=[]
				for xxx in range(0,len(mylist2)):
					mylist.append(str(mylist2[xxx]))
				mylist.sort()

				print(mylist)
		else:
			mylist2 = list(dict.fromkeys(Account))
			mylist=[]
			for xxx in range(0,len(mylist2)):
				mylist.append(str(mylist2[xxx]))
			mylist.sort()

			print(mylist)


		for i in range(1, len(mylist)+1):
			F10TB.cell(row=i+14, column=5).value=int(mylist[i-1])

		for i in range  (  1, len(mylist)+1):
			if(mylist[i-1] in Account):
				F10TB.cell(row=i+14, column=6).value='=VLOOKUP(E{0},PBC_CY!D:E,2,0)'.format(i+14)
			else:
				F10TB.cell(row=i+14, column=6).value='=VLOOKUP(E{0},PBC_PY!D:E,2,0)'.format(i+14)

		for i in range (1,len(mylist)+1):
			x=str(mylist[i-1])
			y=str(x[:4])
			F10TB.cell(row=i+14, column=4).value=str(y)

		for i in range (1,len(mylist)+1):
			x=str(mylist[i-1])
			y=x[:3]
			F10TB.cell(row=i+14, column=3).value=str(y)

		for i in range (1,len(mylist)+1):
			F10TB.cell(row=i+14, column=2).value='=Left(E{0},1)'.format(i+14)


		for i in range(1, len(mylist)+1):
				F10TB.cell(row=i+14, column=1).value='=IF(B{0}<"6","BS",IF(AND(B{0}>"5",B{0}<"8"),"PL","Other Account-Off TB"))'.format(i+14)
		 

		"Calculation"

		for i in range(1, len(mylist)+1):
			if(mylist[i-1] in Account):

				if(int(str(mylist[i-1])[:1])<6):
					F10TB.cell(row=i+14,column=7).value='=SUMIF(PBC_CY!D:D,E{0},PBC_CY!F:F)-SUMIF(PBC_CY!D:D,E{0},PBC_cY!G:G)'.format(i+14)
					F10TB.cell(row=i+14,column=7).number_format='#,##0_);(#,##0)'
			else:
				F10TB.cell(row=i+14,column=7).value='=SUMIF(PBC_PY!D:D,E{0},PBC_PY!F:F)'.format(i+14)
				F10TB.cell(row=i+14,column=7).number_format='#,##0_);(#,##0)'
			if(int(str(mylist[i-1])[:1])==6):
				F10TB.cell(row=i+14,column=7).value='=SUMIF(PBC_PY!D:D,E{0},PBC_PY!F:F)'.format(i+14)
				F10TB.cell(row=i+14,column=7).number_format='#,##0_);(#,##0)'
			if(int(str(mylist[i-1])[:1])==7):
				F10TB.cell(row=i+14,column=7).value='=SUMIF(PBC_PY!D:D,E{0},PBC_PY!F:F)'.format(i+14)
				F10TB.cell(row=i+14,column=7).number_format='#,##0_);(#,##0)'

		for i in range(1, len(mylist)+1):
			F10TB.cell(row=i+14, column=8).value='=SUMIF(PBC_CY!D:D,E{0},PBC_CY!H:H)'.format(i+14)
			F10TB.cell(row=i+14,column=8).number_format='#,##0_);(#,##0)'

		for i in range(1, len(mylist)+1):
			F10TB.cell(row=i+14, column=9).value='=SUMIF(PBC_CY!D:D,E{0},PBC_CY!I:I)'.format(i+14)
			F10TB.cell(row=i+14,column=9).number_format='#,##0_);(#,##0)'

		for i in range(1, len(mylist)+1):
			F10TB.cell(row=i+14, column=10).value='=IF(B{0}<"6",SUMIF(PBC_CY!D:D,E{0},PBC_CY!J:J)-SUMIF(PBC_CY!D:D,E{0},PBC_CY!K:K),IF(AND(B{0}="6",C{0}="609",H{0}>0),-H{0},IF(AND(B{0}="6",H{0}<>I{0}),H{0}-I{0},IF(B{0}="6",H{0},IF(AND(B{0}="7",C{0}="709",I{0}<0),I{0},IF(AND(B{0}="7",C{0}="711"),-$U$6,IF(AND(B{0}="7",C{0}="712"),-$U$8,IF(AND(B{0}="7",H{0}<>I{0}),H{0}-I{0},IF(AND(B{0}="7",I{0}>0),-I{0},IF(AND(B{0}="7",I{0}<0),I{0},0))))))))))'.format(i+14)
			F10TB.cell(row=i+14,column=10).number_format='#,##0_);(#,##0)'


		for i in range(1, len(mylist)+1):
			F10TB.cell(row=i+14, column =11).value='=IF(A{0}="BS",G{0}+H{0}-I{0}-J{0},IF(AND(A{0}="PL",H{0}<>I{0}),H{0}-I{0}-J{0},H{0}-I{0}))'.format(i+14)
			F10TB.cell(row=i+14,column=11).number_format='#,##0_);(#,##0)'
			F10TB.cell(row=i+14, column=11).font=fontRed

		for i in range(1, len(mylist)+1):
			F10TB.cell(row=i+14, column=13).value='=IF(A{0}="BS",J{0}-G{0},"")'.format(i+14)
			F10TB.cell(row=i+14,column=13).number_format='#,##0_);(#,##0)'

		F10TB.cell(row=13, column=12).number_format='#,##0_);(#,##0)'

		for i in range(1, len(mylist)+1):
			F10TB.cell(row=i+14, column=14).value='=IF(A{0}="BS",IF(AND(G{0}=0,J{0}=0),0,IF(AND(G{0}=0,J{0}>0),1,IF(AND(G{0}=0,J{0}<0),-1,IF(AND(J{0}=0,G{0}>0),-1,IF(AND(J{0}=0,G{0}<0),1,J{0}/G{0}-1))))),"")'.format(i+14)
			F10TB.cell(row=i+14,column=14).number_format="0.0%"

		# for i in range(1, len(mylist)+1):
		# 	x=str(mylist[i-1])
		# 	F10TB.cell(row=i+14,column=16).value='=if('+x[:1]+'<6'+",iferror(vlookup(D{0},'BS Mapping'!A:C,3,0),vlookup(C{0},'BS Mapping'!A:C,3,0)),iferror(vlookup(D{0},'PL Mapping'!A:C,3,0),vlookup(C{0},'PL Mapping'!A:C,3,0)))".format(i+14)

		# for i in range(1, len(mylist)+1):
		# 	x=str(mylist[i-1])
		# 	F10TB.cell(row=i+14,column=17).value='=if('+x[:1]+'<6'+",iferror(vlookup(D{0},'BS Mapping'!A:D,4,0),vlookup(C{0},'BS Mapping'!A:D,4,0)),iferror(vlookup(D{0},'PL Mapping'!A:D,4,0),vlookup(C{0},'PL Mapping'!A:D,4,0)))".format(i+14)
		# for i in range(1, len(mylist)+1):
		# 	x=str(mylist[i-1])
		# 	F10TB.cell(row=i+14,column=18).value='=if('+x[:1]+'<6'+",iferror(vlookup(D{0},'BS Mapping'!A:E,5,0),vlookup(C{0},'BS Mapping'!A:E,5,0)),iferror(vlookup(D{0},'PL Mapping'!A:E,5,0),vlookup(C{0},'PL Mapping'!A:E,5,0)))".format(i+14)
		for i in range(1, len(mylist)+1):
			F10TB.cell(row=i+14,column=16).value="=G{0}-SUMIF(PBC_PY!D:D,E{0},PBC_PY!F:F)".format(i+14)
			F10TB.cell(row=i+14,column=16).number_format='#,##0_);(#,##0)'
			F10TB.cell(row=i+14,column=16).font=fontRed
			# F10TB.cell(row=i+14,column=16).value='=if('+x[:1]+'<6,iferror(vlookup('+str(x[0:4])+",'BS Mapping std'!A:E,5,0),vlookup("+str(x[0:3])+",'BS Mapping std'!A:E,5,0)),iferror(vlookup("+str(x[0:4])+",'PL mapping Std'!A:E,5,0),vlookup("+str(x[0:3])+",'PL mapping Std'!A:E,5,0))"
		"Closing 711"

		F10TB.cell(row=1, column=14).value="Closing 711"
		F10TB.cell(row=1, column=14).font=font2
		F10TB.cell(row=1, column=14).alignment=Alignment(horizontal='right')

		F10TB.cell(row=1, column=14).value="Acc."
		F10TB.cell(row=1, column=15).value="OB"
		F10TB.cell(row=1, column=16).value="CB"
		F10TB.cell(row=1, column=17).value="VAR"

		F10TB.cell(row=2, column=14).value="331"
		F10TB.cell(row=3, column=14).value="341"
		F10TB.cell(row=4, column=14).value="345"
		F10TB.cell(row=5, column=14).value="348"
		F10TB.cell(row=6, column=14).value="Total:"
		F10TB.cell(row=6, column=14).font=font
		F10TB.cell(row=6, column=14).alignment=Alignment(horizontal='right' )

		F10TB.cell(row=8, column=14).value="332" 
		F10TB.cell(row=8, column=14).font=font
		#wrerwe

		for i in range(14,18):
			F10TB.cell(row=1, column=i).font=font2
			F10TB.cell(row=1, column=i).fill=blueFill
			F10TB.cell(row=1, column=i).border=doubleborder
			F10TB.cell(row=1, column=i).alignment=Alignment(horizontal='left')

		for i in range (2,9):
			F10TB.cell(row=i, column=14).font=font
			F10TB.cell(row=i, column=14).alignment=Alignment(horizontal='right')

		for i in range (14,18):
			F10TB.cell(row=5, column=i).border=doubleborder

		for i in range (2, 6):
			F10TB.cell(row=i, column=15).value='=SUMIF(C:C,N{0},G:G)'.format(i)
			F10TB.cell(row=i, column=16).value='=SUMIF(C:C,N{0},J:J)'.format(i)
			F10TB.cell(row=i, column=17).value='=P{0}-O{0}'.format(i)

		F10TB.cell(row=6, column=15).value='=SUM(O2:O5)'
		F10TB.cell(row=6, column=16).value='=SUM(P2:P5)'
		F10TB.cell(row=6, column=17).value='=P6-O6'
		F10TB.cell(row=6, column=17).font=font2
		F10TB.cell(row=6, column=17).fill=blueFill

		for i in range (15,18):
			F10TB.cell(row=6, column=i).font=font2
		F10TB.cell(row=14,column=16).fill=blueFill
		F10TB.cell(row=14,column=16).font=font2

		for i in range(2,14):
			for j in range(15,18):  
				F10TB.cell(row=i, column=j).number_format='#,##0_);(#,##0)'

		F10TB.cell(row=7, column=14).value="Closing 712"
		F10TB.cell(row=7, column=14).font=font2
		F10TB.cell(row=7, column=14).alignment=Alignment(horizontal='right')

		F10TB.cell(row=8, column=15).value='=SUMIF(C:C,N8,G:G)'
		F10TB.cell(row=8, column=16).value='=SUMIF(C:C,N8,J:J)'
		F10TB.cell(row=8, column=17).value='=P8-O8'
		F10TB.cell(row=8, column=17).fill=blueFill
		F10TB.cell(row=8, column=17).font=font2

		for i in range (6,10):
			F10TB.cell(row=11, column=i).border=doubleborder

		F10TB.auto_filter.ref = 'A14:P14'

		c = F10TB['B15']
		F10TB.freeze_panes = c


		x
		"Adjust Column Width"

		for col in F10TB.columns:
			max_length = 0
			for cell in col:
				if cell.coordinate in F10TB.merged_cells:
					continue
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(cell.value)
				except:
					pass
			adjusted_width=(max_length-5)


		listanoua=['F','G','H','I','J','K','M','N','O','L']
		for column in ascii_uppercase:
			for i in listanoua:
				if (column==i):
					F10TB.column_dimensions[column].width =15

		listanoua2=['A']
		for column in ascii_uppercase:
			for i in listanoua2:
				if (column==i):
					F10TB.column_dimensions[column].width = 10
#a

		
		file_path=os.path.join(folderpath, "F100 Trial Balance.xlsx")
		myorder=[3,2,1]
		output._sheets =[output._sheets[i] for i in myorder]
		output.save(folderpath+"/Trial Balance.xlsx")
		return send_from_directory(folderpath,"Trial Balance.xlsx",as_attachment=True)

		# print(text)

@app.route('/VAT/Instructions', methods=['GET'])
def downloadVAT():
		filepath = "/home/fsbot/storage/vat"
 
		return send_from_directory(filepath,"Instructions - VAT.docx", as_attachment=True)
@app.route('/VAT/GTbJY47MKf1oajfEqntaRFSt8fw')
def my_formVAT():
	return render_template('VAT.html')
@app.route('/VAT/GTbJY47MKf1oajfEqntaRFSt8fw', methods=['POST', 'GET'])
def my_form_post():
	yearEnd1 = datetime.datetime.strptime(
		request.form['yearEnd'],
		'%Y-%m-%d')
	preparedBy1 = request.form['preparedBy']
	clientname1 = request.form['client']   
	datePrepared1 = datetime.datetime.strptime(
		request.form['preparedDate'],
		'%Y-%m-%d')
	refference1 = request.form['reff']
	denis=datetime.datetime.now()
#
# 	# yearEnd = str(request.form['yearEnd'])
# 	# processed_text = client.upper()
# 	# fisier=request.files.get('monthlyTB')
	if request.method == 'POST':
		def getAttachments(reader):
			catalog = reader.trailer["/Root"]
			fileNames = catalog['/Names']['/EmbeddedFiles']['/Names']
			attachments = {}
			for f in fileNames:
				if isinstance(f, str):
					name = f
					dataIndex = fileNames.index(f) + 1
					fDict = fileNames[dataIndex].getObject()
					fData = fDict['/EF']['/F'].getData()  
					attachments[name] = fData

			return attachments
		# file_TemplateXML = request.files('vatXML')
		file_TemplateXML = request.files.getlist('vatXML')      
		file_TB = request.files["TB"]

		# for i in file_TemplateXML:
		# 	i.save(secure_filename(i.filename))

		# fonts and colors
		ft1 = Font(name='Arial', size=10, bold=True)
		f_testname = Font(name='Arial', size=15, color='614C77', bold=True)
		f_info = Font(name='Arial', size=10, color='614C77', bold=True)
		cap_tabel = Font(name='Arial', size=10, color="FFFFFF", bold=True)
		cap_tabel_color_PBC = PatternFill(start_color='808080', end_color='808080', fill_type='solid')  # grey
		cap_tabel_color_GT = PatternFill(start_color='00AEAC', end_color='00AEAC', fill_type='solid')  # indigo #B1A0C7
		fprocentaj = Font(name='Arial', size=10, color="FF0000", bold=True)
		font_worksheet = Font(name='Arial', size=10)
		check_font = Font(name='Arial', size=10, color="FF0000", bold=True)
		check_font_1 = Font(name='Arial', size=10, color="FF0000", bold=False)
		cap_tabel_color_GT_movdeschis = PatternFill(start_color='00AEAC', end_color='00AEAC', fill_type='solid')
		cap_tabel_color_GT_movinchis = PatternFill(start_color='3BBCCA', end_color='3BBCCA', fill_type='solid')
		blue_bold_font = Font(name='Arial', size=10, color="0070C0", bold=True)
		blue_thin_font = Font(name='Arial', size=10, color="0070C0", bold=False)
		# TB_font = Font(name='Arial', size=10, color='0070C0', bold=True)

		thin = Side(border_style='thin', color='000000')
		border = Border(left=thin, right=thin, top=thin, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_left = Border(left=thin, right=None, top=thin, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_right = Border(left=None, right=thin, top=thin, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_centered = Border(left=None, right=None, top=thin, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_upperleft = Border(left=thin, top=thin)

		thin = Side(border_style='thin', color='000000')
		border_lowerleft = Border(left=thin, right=None, top=None, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_upperright = Border(right=thin, top=thin)

		thin = Side(border_style='thin', color='000000')
		border_lowerright = Border(right=thin, bottom=thin)

		thin = Side(border_style='thin', color='000000')
		border_left1 = Border(left=thin)

		thin = Side(border_style='thin', color='000000')
		border_right1 = Border(right=thin)

		thin = Side(border_style='thin', color='000000')
		border_top = Border(top=thin)

		thin = Side(border_style='thin', color='000000')
		border_bottom = Border(bottom=thin)

		# app.mainloop()

		wb = openpyxl.Workbook()
		ws = wb.active

		Sheet1 = wb.create_sheet("Lead Schedule")
		Sheet1.sheet_view.showGridLines = False

		Sheet1.cell(row = 1, column = 2).value ="Client Name:"
		Sheet1.cell(row = 2, column = 2).value ="Period End:"
		Sheet1.cell(row = 1, column = 15).value ="Prepared By:"
		Sheet1.cell(row = 2, column = 15).value ="Date:"
		Sheet1.cell(row = 3, column = 15).value ="Reviewed by:"
		Sheet1.cell(row = 5, column = 2).value ="Account Reconciliation & SA "

		Sheet1.cell(row=7, column=2).value = "Class"
		Sheet1.cell(row=7, column=3).value = "Synt 3"
		Sheet1.cell(row=7, column=4).value = "Synt 4"
		Sheet1.cell(row=7, column=5).value = "Account"
		Sheet1.cell(row=7, column=6).value = "Descriere"
		Sheet1.cell(row=7, column=7).value = "OB"
		Sheet1.cell(row=7, column=8).value = "DR"
		Sheet1.cell(row=7, column=9).value = "CR"
		Sheet1.cell(row=7, column=10).value = "CB"
		Sheet1.cell(row=7, column=11).value = "Variation"
		Sheet1.cell(row=7, column=12).value = "Variation %"

		Sheet1.cell(row=7, column=2).value = "Class"
		Sheet1.cell(row=7, column=3).value = "Synt 3"
		Sheet1.cell(row=7, column=4).value = "Synt 4"
		Sheet1.cell(row=7, column=5).value = "Account"
		Sheet1.cell(row=7, column=6).value = "Descriere"

		Sheet1.cell(row=8, column=2).value ="4"
		Sheet1.cell(row=9, column=2).value ="4"
		Sheet1.cell(row=10, column=2).value ="4"
		Sheet1.cell(row=11, column=2).value ="4"

		Sheet1.cell(row=8, column=3).value ="442"
		Sheet1.cell(row=9, column=3).value ="442"
		Sheet1.cell(row=10, column=3).value ="442"
		Sheet1.cell(row=11, column=3).value ="442"

		Sheet1.cell(row=8, column=4).value ="4423"
		Sheet1.cell(row=9, column=4).value ="4426"
		Sheet1.cell(row=10, column=4).value ="4427"
		Sheet1.cell(row=11, column=4).value ="4428"

		Sheet1.cell(row=8, column=5).value ="Taxa pe valoarea adaugata de plata"
		Sheet1.cell(row=9, column=5).value ="Taxa pe valoarea adaugata deductibil"
		Sheet1.cell(row=10, column=5).value ="Taxa pe valoarea adaugata colectat"
		Sheet1.cell(row=11, column=5).value ="Taxa pe valoarea adaugata neexigibila"

		Sheet1.cell(row=12, column=6).value = "Total"

		#content
		Sheet1.cell(row=8, column=7).value = "=SUMIF('TB Robot'!B:B,4423,'TB Robot'!E:E)".format(16)
		Sheet1.cell(row=9, column=7).value = "=SUMIF('TB Robot'!B:B,4426,'TB Robot'!E:E)".format(16)
		Sheet1.cell(row=10, column=7).value = "=SUMIF('TB Robot'!B:B,4427,'TB Robot'!E:E)".format(16)
		Sheet1.cell(row=11, column=7).value = "=SUMIF('TB Robot'!B:B,4428,'TB Robot'!E:E)".format(16)
		Sheet1.cell(row=12, column=7).value = "=SUM(G8:G11)".format(16)

		Sheet1.cell(row=8, column=8).value = "=SUMIF('TB Robot'!B:B,4423,'TB Robot'!F:F)".format(16)
		Sheet1.cell(row=9, column=8).value = "=SUMIF('TB Robot'!B:B,4426,'TB Robot'!F:F)".format(16)
		Sheet1.cell(row=10, column=8).value = "=SUMIF('TB Robot'!B:B,4427,'TB Robot'!F:F)".format(16)
		Sheet1.cell(row=11, column=8).value = "=SUMIF('TB Robot'!B:B,4428,'TB Robot'!F:F)".format(16)
		Sheet1.cell(row=12, column=8).value = "=SUM(H8:H11)".format(16)

		Sheet1.cell(row=8, column=9).value = "=SUMIF('TB Robot'!B:B,4423,'TB Robot'!G:G)".format(16)
		Sheet1.cell(row=9, column=9).value = "=SUMIF('TB Robot'!B:B,4426,'TB Robot'!G:G)".format(16)
		Sheet1.cell(row=10, column=9).value = "=SUMIF('TB Robot'!B:B,4427,'TB Robot'!G:G)".format(16)
		Sheet1.cell(row=11, column=9).value = "=SUMIF('TB Robot'!B:B,4428,'TB Robot'!G:G)".format(16)
		Sheet1.cell(row=12, column=9).value = "=SUM(I8:I11)".format(16)

		Sheet1.cell(row=8, column=10).value = "=SUMIF('TB Robot'!B:B,4423,'TB Robot'!H:H)".format(16)
		Sheet1.cell(row=9, column=10).value = "=SUMIF('TB Robot'!B:B,4426,'TB Robot'!H:H)".format(16)
		Sheet1.cell(row=10, column=10).value = "=SUMIF('TB Robot'!B:B,4427,'TB Robot'!H:H)".format(16)
		Sheet1.cell(row=11, column=10).value = "=SUMIF('TB Robot'!B:B,4428,'TB Robot'!H:H)".format(16)
		Sheet1.cell(row=12, column=10).value = "=SUM(J8:J11)".format(16)

		Sheet1.cell(row=8, column=11).value = "=J8-G8".format(16)
		Sheet1.cell(row=9, column=11).value = "=J9-G9".format(16)
		Sheet1.cell(row=10, column=11).value = "=J10-G10".format(16)
		Sheet1.cell(row=11, column=11).value = "=J11-G11".format(16)
		Sheet1.cell(row=12, column=11).value = "=SUM(K8:K11)".format(16)

		Sheet1.cell(row=8, column=12).value = "=IFERROR(K8/G8,1)".format(16)
		Sheet1.cell(row=9, column=12).value = "=IFERROR(K9/G9,1)".format(16)
		Sheet1.cell(row=10, column=12).value = "=IFERROR(K10/G10,1)".format(16)
		Sheet1.cell(row=11, column=12).value = "=IFERROR(K11/G11,1)".format(16)
		Sheet1.cell(row=12, column=12).value = "=IFERROR(K12/G12,1)".format(16)

		Sheet1.cell(row = 1, column = 3).value =clientname1
		Sheet1.cell(row = 1, column = 3).value =clientname1
		Sheet1.cell(row = 2, column = 3).value =yearEnd1
		Sheet1.cell(row = 2, column = 3).number_format = 'mm/dd/yyyy'
		Sheet1.cell(row = 1, column = 16).value =preparedBy1
		Sheet1.cell(row = 2, column = 16).value =datePrepared1
		Sheet1.cell(row = 2, column = 16).number_format = 'mm/dd/yyyy'

		#design
		Sheet1.cell(row = 1, column = 2).font =ft1
		Sheet1.cell(row = 2, column = 2).font =ft1
		Sheet1.cell(row = 1, column = 15).font =ft1
		Sheet1.cell(row = 2, column = 15).font =ft1
		Sheet1.cell(row = 3, column = 15).font =ft1
		Sheet1.cell(row = 5, column = 2).font = f_testname

		for row in Sheet1['B7:L7']:
			for cell in row:
				cell.fill = cap_tabel_color_GT

		for row in Sheet1['B7:L7']:
			for cell in row:
				cell.font = cap_tabel

		for row in Sheet1['B8:L11']:
			for cell in row:
				cell.font = font_worksheet

		for row in Sheet1['G8:L12']:
			for cell in row:
				cell.number_format = '#,##0_);(#,##0)'

		for row in Sheet1['B7:L11']:
			for cell in row:
				cell.border = border

		Sheet1.column_dimensions['F'].hidden = True
		Sheet1.column_dimensions['F'].width = 40
		Sheet1.column_dimensions['G'].width = 10
		Sheet1.column_dimensions['H'].width = 10
		Sheet1.column_dimensions['I'].width = 10
		Sheet1.column_dimensions['J'].width = 10
		Sheet1.column_dimensions['K'].width = 14
		Sheet1.column_dimensions['L'].width = 14

		Sheet2 = wb.create_sheet("VAT Test")
		Sheet2.sheet_view.showGridLines = False

		Sheet2.cell(row = 1, column = 2).value ="Client Name:"
		Sheet2.cell(row = 2, column = 2).value ="Period End:"
		Sheet2.cell(row = 1, column = 10).value ="Prepared By:"
		Sheet2.cell(row = 2, column = 10).value ="Date:"
		Sheet2.cell(row = 3, column = 10).value ="Reviewed by:"

		Sheet2.cell(row = 5, column = 2).value ="VAT Reconciliation Summary"
		Sheet2.cell(row = 7, column = 2).value ="Procedures"
		Sheet2.cell(row = 8, column = 2).value ="We have obtained the monthly VAT statement prepared for Local Budget (no. 300)."
		Sheet2.cell(row = 9, column = 2).value ="We have reconciled the VAT from the VAT statement and VAT from the sales and acquisition journals with the value from trial balance."

		Sheet2.cell(row = 11, column = 2).value ="Work Done"
		Sheet2.cell(row = 12, column = 2).value ="Please see below:"

		Sheet2.cell(row = 16, column = 2).value = "Month"
		Sheet2.cell(row = 16, column = 3).value = "Refference"
		Sheet2.cell(row = 16, column = 5).value = "Sales 19% as per VAT Statement"
		Sheet2.cell(row = 16, column = 6).value = "VAT collected 19% as per Nexia"
		Sheet2.cell(row = 16, column = 7).value = "Sales 9% as per VAT Statement"
		Sheet2.cell(row = 16, column = 8).value = "VAT collected 9% as per Nexia"
		Sheet2.cell(row = 16, column = 9).value = "Sales 5% as per VAT Statement"
		Sheet2.cell(row = 16, column = 10).value = "VAT collected 5% as per Nexia"
		Sheet2.cell(row = 16, column = 11).value = "Sales w/o VAT as per VAT Statement"
		Sheet2.cell(row = 16, column = 12).value = "Reverse Taxation as per VAT Statement"
		Sheet2.cell(row = 16, column = 13).value = "VAT collected for Reverse Taxation as per Nexia (19%)"
		Sheet2.cell(row = 16, column = 14).value = "Regularisation as per VAT Statement"
		Sheet2.cell(row = 16, column = 15).value = "VAT collected for Regularization as per Nexia (19%/9%/5%)"
		Sheet2.cell(row = 16, column = 17).value = "Total sales as per VAT Statement"
		Sheet2.cell(row = 16, column = 18).value = "Total sales as per Nexia Calculation"
		Sheet2.cell(row = 16, column = 19).value = "Difference on Total Sales"
		Sheet2.cell(row = 16, column = 21).value ="Total VAT Collected as per VAT Statement"
		Sheet2.cell(row = 16, column = 22).value ="Total VAT collected as per Nexia"
		Sheet2.cell(row = 16, column = 23).value ="Difference on VAT Collected Company vs Nexia"
		Sheet2.cell(row = 16, column = 25).value ="Purchases  19% as per VAT Statement"
		Sheet2.cell(row = 16, column = 26).value ="VAT deductible  as per Nexia"
		Sheet2.cell(row = 16, column = 27).value ="Purchases 9% as per VAT Statement"
		Sheet2.cell(row = 16, column = 28).value ="VAT 9% as per Nexia"
		Sheet2.cell(row = 16, column = 29).value ="Purchases 5% as per VAT Statement"
		Sheet2.cell(row = 16, column = 30).value ="VAT 5% as per Nexia"
		Sheet2.cell(row = 16, column = 31).value ="Reverse Taxation as per VAT Statement"
		Sheet2.cell(row = 16, column = 32).value ="VAT deducted for Reverse Taxation as per Nexia (19%)"
		Sheet2.cell(row = 16, column = 33).value ="Total Purchases w/o VAT as VAT per Statement"
		Sheet2.cell(row = 16, column = 34).value ="Regularisation as per VAT Statement"
		Sheet2.cell(row = 16, column = 35).value ="VAT deducted for Regularization as per Nexia (19%/9%/5%)"
		Sheet2.cell(row = 16, column = 37).value ="Total Purchases with VAT as per VAT Statement"
		Sheet2.cell(row = 16, column = 38).value ="Total Purchases with VAT as per Nexia"
		Sheet2.cell(row = 16, column = 39).value ="Difference on Total Purchases with VAT Company vs Nexia"
		Sheet2.cell(row = 16, column = 41).value ="Total VAT deductible as per VAT Statement"
		Sheet2.cell(row = 16, column = 42).value ="Total VAT deductible as per Nexia"
		Sheet2.cell(row = 16, column = 43).value ="Difference on Total VAT deductible Company vs Nexia"
		Sheet2.cell(row = 16, column = 45).value ="VAT (due)/to be recovered per client"
		Sheet2.cell(row = 16, column = 46).value ="VAT (due)/to be recovered per Nexia"
		Sheet2.cell(row = 16, column = 47).value ="Difference Company vs Nexia"
		Sheet2.cell(row = 16, column = 48).value ="VAT deducted as per statement"
		Sheet2.cell(row = 16, column = 49).value ="VAT deductible vs VAT deducted"
		Sheet2.cell(row = 16, column = 50).value ="Remaining difference statement vs Nexia"
		Sheet2.cell(row = 16, column = 52).value ="Payment order/ Request for reimbursement or compensation no/date during the month"
		Sheet2.cell(row = 16, column = 53).value ="Amount "
		Sheet2.cell(row = 16, column = 54).value ="VAT balance"

		Sheet2.cell(row=34, column=19).value ="Credit Movement of account #4427 as per T/B"
		Sheet2.cell(row=35, column=19).value ="Difference"

		Sheet2.cell(row=34, column=39).value = "Debit Movement of account #4426 as per T/B"
		Sheet2.cell(row=35, column=39).value = "Difference Detail vs TB"

		Sheet2.cell(row=37, column=39).value = "Debit Movement of account #4423 as per T/B"
		Sheet2.cell(row=38, column=39).value = "Difference Detail vs TB"

		Sheet2.cell(row=43, column=6).value = "Total Sales excluding Reverse Taxation as per VAT Statement"
		Sheet2.cell(row=44, column=6).value = "Total Sales as per TB"
		Sheet2.cell(row=45, column=6).value = "Mvm 70x"
		Sheet2.cell(row=46, column=6).value = "Variation 419 (Cr-Db)"
		Sheet2.cell(row=47, column=6).value = "Variation 418 -(Db-Cr)"
		Sheet2.cell(row=48, column=6).value = "Variation 472 (Cr-Db)"
		Sheet2.cell(row=49, column=6).value = "Mvm 758X"
		Sheet2.cell(row=50, column=6).value = "Total"
		Sheet2.cell(row=52, column=5).value ="Difference"
		Sheet2.cell(row=53, column=5).value ="Maximum impact on VAT"

		Sheet2.cell(row=58, column=6).value = "Total Purchases as per VAT Statement"
		Sheet2.cell(row=59, column=6).value = "Total Purchases as per TB"
		Sheet2.cell(row=60, column=6).value = "Mvm 60x "
		Sheet2.cell(row=61, column=6).value = "Variation #3xx other than #39x"
		Sheet2.cell(row=62, column=6).value = "FA acqusitions"
		Sheet2.cell(row=63, column=6).value = "Mvm 61X"
		Sheet2.cell(row=64, column=6).value = "Mvm 62x"
		Sheet2.cell(row=65, column=6).value = "Mvm 658X"
		Sheet2.cell(row=66, column=6).value = "Total"

		Sheet2.cell(row=68, column=5).value = "Difference"
		Sheet2.cell(row=69, column=5).value = "Maximum impact on VAT"

		#content
		Sheet2.cell(row = 1, column = 3).value =clientname1
		Sheet2.cell(row = 2, column = 3).value =yearEnd1
		Sheet2.cell(row = 2, column = 3).number_format='mm/dd/yyyy'
		Sheet2.cell(row = 1, column = 11).value =preparedBy1
		Sheet2.cell(row = 2, column = 11).value =datePrepared1
		Sheet2.cell(row = 2, column = 11).number_format='mm/dd/yyyy'

		#retinem variabilele din XML
		files=list(file_TemplateXML)
		def changeWord(word):
			for letter in word:
				if letter == "b'":
					word = word.replace(letter,'')
			return word
		nr=0
		abc=[]
		now=datetime.datetime.now()
		for i in files:
			nr=nr+1
			reader = PyPDF2.PdfFileReader(i)
			dictionary = getAttachments(reader)
			a=str(*dictionary.values())
			b=a.replace('\\n','').replace("b'","").replace("></declaratie300>'","></declaratie300>").replace('encoding="utf-8"', "").replace('''"/>''', "</declaratie300>").replace("</declaratie300>'", '"></declaratie300>').replace('\\r', '')
			# print(b)
			f=open("/home/fsbot/storage/vat"+str(nr)+str(now.year)+str(now.month)+str(now.day)+str(now.hour)+str(now.minute)+str(now.second)+".xml","w").write(b)
			abc.append("/home/fsbot/storage/vat"+str(nr)+str(now.year)+str(now.month)+str(now.day)+str(now.hour)+str(now.minute)+str(now.second)+".xml")
			
		fisiere=list(abc)
		for v in range(0,len(fisiere)):
			tree = ET.parse(fisiere[v])
			root = tree.getroot()
			month = int(root.attrib["luna"])
			print(month)
			for j in range(1, month+1):
				Sheet2.cell(row=16 + j, column=2).value = j
				# Sheet2.cell(row=36 + j, column=2).value = j
				# Sheet2.cell(row=54 + j, column=2).value = j
			# a=0
			# try:
			# 	a = a + int(root.attrib['R9_1'])
			# except:
			# 	print("n a mers")
			#
			# try:
			# 	a = a + int(root.attrib['R17_1'])
			# except:
			# 	print("n a mers 2")

			try:
				Sheet2.cell(row=16 + month, column=5).value = int(root.attrib['R9_1'])
			except:
				Sheet2.cell(row=16 + month, column=5).value = 0


			# Sheet2.cell(row=15+i, column=2).value = int(root.attrib["luna"])

			Sheet2.cell(row = 16 + month, column = 6).value = "=(E{0}*0.19)*1".format(16 + month)


			try:
				Sheet2.cell(row = 16 + month, column = 7).value = int(root.attrib['R10_1'])
			except:
				Sheet2.cell(row = 16 + month, column = 7).value = 0

			Sheet2.cell(row = 16 + month, column = 8).value = "=(G{0}*0.09)*1".format(16 + month)

			try:
				Sheet2.cell(row = 16 + month, column = 9).value = int(root.attrib['R11_1'])
			except:
				Sheet2.cell(row = 16 + month, column = 9).value = 0

			Sheet2.cell(row = 16 + month, column = 10).value = "=(I{0}*0.05)*1".format(16 + month)

			x=0
			try:
				x = x + int(root.attrib['R1_1'])
			except:
				print('b')

			try:
				x = x + int(root.attrib['R2_1'])
			except:
				print('b')

			try:
				x = x + int(root.attrib['R3_1'])
			except:
				print('b')

			try:
				x = x + int(root.attrib['R4_1'])
			except:
				print('b')

			try:
				x = x + int(root.attrib['R13_1'])
			except:
				print('b')

			try:
				x = x + int(root.attrib['R14_1'])
			except:
				print('b')

			try:
				x = x + int(root.attrib['R15_1'])
			except:
				print('b')

			Sheet2.cell(row = 16 + month, column = 11).value = x

			x = 0
			# try:
			# 	x = x + int(root.attrib['R5_1'])
			# except:
			# 	print('5_1')

			try:
				x = x + int(root.attrib['R6_1'])
			except:
				print('6_1')

			try:
				x = x + int(root.attrib['R7_1'])
			except:
				print('7_1')

			try:
				x = x + int(root.attrib['R8_1'])
			except:
				print('8_1')

			try:
				x = x + int(root.attrib['R12_1'])
			except:
				print('12_1')
			Sheet2.cell(row = 16 + month, column = 12).value = x
			Sheet2.cell(row = 16 + month, column = 13).value = "=(L{0}*0.19)*1".format(16 + month)

			b= 0
			try:
				b = b + int(root.attrib['R16_1'])
			except:
				print("nu merge 16_1")

			try:
				b = b + int(root.attrib['R18_1'])
			except:
				print("nu merge 18_1")

			try:
				Sheet2.cell(row = 16 + month, column = 14).value = b
			except:
				Sheet2.cell(row = 16 + month, column = 14).value = 0

			Sheet2.cell(row = 16 + month, column = 15).value = "=(N{0}*0.19)*1".format(16 + month)

			try:
				Sheet2.cell(row = 16 + month, column = 17).value = int(root.attrib['R17_1']) #C1
			except:
				Sheet2.cell(row = 16 + month, column = 17).value = int(root.attrib['R19_1'])

			Sheet2.cell(row = 16 + month, column = 18).value = "=(E{0}+G{0}+K{0}+L{0}+N{0}+I{0})*1".format(16 + month)
			Sheet2.cell(row = 16 + month, column = 19).value = "=Q{0}-R{0}".format(16 + month)

			try:
				Sheet2.cell(row = 16 + month, column = 21).value = int(root.attrib['R17_2']) #C2
			except:
				Sheet2.cell(row = 16 + month, column = 21).value = 0

			Sheet2.cell(row = 16 + month, column = 22).value = "=(F{0}+J{0}+H{0}+M{0}+O{0})*1".format(16 + month)
			Sheet2.cell(row = 16 + month, column = 23).value = "=(U{0}-V{0})*1".format(16 + month)

			try:
				Sheet2.cell(row = 16 + month, column = 25).value = int(root.attrib['R22_1'])
			except:
				Sheet2.cell(row = 16 + month, column = 25).value = 0
			Sheet2.cell(row = 16 + month, column = 26).value = "=(Y{0}*0.19)*1".format(16 + month)

			try:
				Sheet2.cell(row = 16 + month, column = 27).value = int(root.attrib['R23_1'])
			except:
				Sheet2.cell(row = 16 + month, column = 27).value = 0
			Sheet2.cell(row = 16 + month, column = 28).value = "=(AA{0}*0.09)*1".format(16 + month)

			try:
				Sheet2.cell(row = 16 + month, column = 29).value = int(root.attrib['R24_1'])
			except:
				Sheet2.cell(row = 16 + month, column = 29).value = 0
			Sheet2.cell(row = 16 + month, column = 30).value = "=(AC{0}*0.05)*1".format(16 + month)

			x = 0
			try:
				x = x + int(root.attrib['R18_1'])
			except:
				print('a')

			try:
				x = x + int(root.attrib['R19_1'])
			except:
				print('a')

			try:
				x = x + int(root.attrib['R20_1'])
			except:
				print('a')
			try:
				x = x + int(root.attrib['R21_1'])
			except:
				print('a')

			try:
				x = x + int(root.attrib['R25_1'])
			except:
				print('a')

			Sheet2.cell(row = 16 + month, column = 31).value = x
			Sheet2.cell(row = 16 + month, column = 32).value = "=(AE{0}*0.19)*1".format(16 + month)

			try:
				Sheet2.cell(row = 16 + month, column = 33).value = int(root.attrib['R26_1'])
			except:
				Sheet2.cell(row = 16 + month, column = 33).value = 0

			try:
				Sheet2.cell(row = 16 + month, column = 34).value = int(root.attrib['R30_2'])
			except:
				Sheet2.cell(row = 16 + month, column = 34).value = 0
			try:
				Sheet2.cell(row=16 + month, column=35).value = int(root.attrib['R30_2'])
			except:
				Sheet2.cell(row=16 + month, column=35).value = 0

			try:
				Sheet2.cell(row = 16 + month, column = 37).value = int(root.attrib['R27_1']) #C1
			except:
				Sheet2.cell(row = 16 + month, column = 37).value = 0

			Sheet2.cell(row = 16 + month, column = 38).value = "=(Y{0}+AA{0}+AE{0}+AC{0})*1".format(16 + month)
			Sheet2.cell(row = 16 + month, column = 39).value = "=(ABS(AK{0}-AL{0}))*1".format(16 + month)

			try:
				Sheet2.cell(row = 16 + month, column = 41).value = int(root.attrib['R27_2']) #C2
			except:
				Sheet2.cell(row = 16 + month, column = 41).value = 0

			Sheet2.cell(row = 16 + month, column = 42).value = "=(Z{0}+AB{0}+AF{0}+AD{0})*1".format(16 + month)
			Sheet2.cell(row = 16 + month, column = 43).value = "=(AO{0}-AP{0})*1".format(16 + month)

			# try:
			# 	Sheet2.cell(row = 16 + month, column = 45).value = int(root.attrib['R33_2'])
			# except:
			# 	if int(root.attrib['R33_2']) == 0:
			# 		Sheet2.cell(row=16 + month, column=45).value = int(root.attrib['R34_2'])
			# 	else:
			# 		Sheet2.cell(row=16 + month, column=45).value = 0

			# Sheet2.cell(row = 16 + month, column = 45).value = int(root.attrib['R33_2'])
			# if int(root.attrib['R33_2']) == 0:
			# 	Sheet2.cell(row=16 + month, column=45).value = int(-int(root.attrib['R34_2']))
			# else:
			# 	Sheet2.cell(row=16 + month, column=45).value = int(root.attrib['R33_2'])

			try:
				Sheet2.cell(row=16 + month, column=45).value = int(-int(root.attrib['R34_2']))
			except:
				Sheet2.cell(row=16 + month, column=45).value = int(root.attrib['R33_2'])

			Sheet2.cell(row = 16 + month, column = 46).value = "=(AP{0}-V{0}+AI{0})*1".format(16 + month)
			Sheet2.cell(row = 16 + month, column = 47).value = "=(AS{0}-AT{0})*1".format(16 + month)

			try:
				Sheet2.cell(row=16 + month, column=48).value = int(root.attrib['R28_2'])
			except:
				Sheet2.cell(row=16 + month, column=48).value = 0

			Sheet2.cell(row=16 + month, column=49).value ="=AP{0}-AV{0}".format(16 + month)
			Sheet2.cell(row=16 + month, column=50).value = "=AU{0}+AW{0}".format(16 + month)

		#calculam totalurile
			Sheet2.cell(row = 30, column = 5).value = "=sum(E17:E28)".format(16 + month)
			Sheet2.cell(row = 30, column = 6).value = "=sum(F17:F28)".format(16 + month)
			Sheet2.cell(row = 30, column = 7).value = "=sum(G17:G28)".format(16 + month)
			Sheet2.cell(row = 30, column = 8).value = "=sum(H17:H28)".format(16 + month)
			Sheet2.cell(row = 30, column = 9).value = "=sum(I17:I28)".format(16 + month)
			Sheet2.cell(row = 30, column = 10).value = "=sum(J17:J28)".format(16 + month)
			Sheet2.cell(row = 30, column = 11).value = "=sum(K17:K28)".format(16 + month)
			Sheet2.cell(row = 30, column = 12).value = "=sum(L17:L28)".format(16 + month)
			Sheet2.cell(row = 30, column = 13).value = "=sum(M17:M28)".format(16 + month)
			Sheet2.cell(row = 30, column = 14).value = "=sum(N17:N28)".format(16 + month)
			Sheet2.cell(row = 30, column = 15).value = "=sum(O17:O28)".format(16 + month)
			Sheet2.cell(row = 30, column = 17).value = "=sum(Q17:Q28)".format(16 + month)
			Sheet2.cell(row = 30, column = 18).value = "=sum(R17:R28)".format(16 + month)
			Sheet2.cell(row = 30, column = 19).value = "=sum(S17:S28)".format(16 + month)
			Sheet2.cell(row = 30, column = 21).value = "=sum(U17:U28)".format(16 + month)
			Sheet2.cell(row = 30, column = 22).value = "=sum(V17:V28)".format(16 + month)
			Sheet2.cell(row = 30, column = 23).value = "=sum(W17:W28)".format(16 + month)
			Sheet2.cell(row = 30, column = 25).value = "=sum(Y17:Y28)".format(16 + month)
			Sheet2.cell(row = 30, column = 26).value = "=sum(Z17:Z28)".format(16 + month)
			Sheet2.cell(row = 30, column = 27).value = "=sum(AA17:AA28)".format(16 + month)
			Sheet2.cell(row = 30, column = 28).value = "=sum(AB17:AB28)".format(16 + month)
			Sheet2.cell(row = 30, column = 29).value = "=sum(AC17:AC28)".format(16 + month)
			Sheet2.cell(row = 30, column = 30).value = "=sum(AD17:AD28)".format(16 + month)
			Sheet2.cell(row = 30, column = 31).value = "=sum(AE17:AE28)".format(16 + month)
			Sheet2.cell(row = 30, column = 32).value = "=sum(AF17:AF28)".format(16 + month)
			Sheet2.cell(row = 30, column = 33).value = "=sum(AG17:AG28)".format(16 + month)
			Sheet2.cell(row = 30, column = 34).value = "=sum(AH17:AH28)".format(16 + month)
			Sheet2.cell(row = 30, column = 35).value = "=sum(AI17:AI28)".format(16 + month)
			Sheet2.cell(row = 30, column = 37).value = "=sum(AK17:AK28)".format(16 + month)
			Sheet2.cell(row = 30, column = 38).value = "=sum(AL17:AL28)".format(16 + month)
			Sheet2.cell(row = 30, column = 39).value = "=sum(AM17:AM28)".format(16 + month)
			Sheet2.cell(row = 30, column = 41).value = "=sum(AO17:AO28)".format(16 + month)
			Sheet2.cell(row = 30, column = 42).value = "=sum(AP17:AP28)".format(16 + month)
			Sheet2.cell(row = 30, column = 43).value = "=sum(AQ17:AQ28)".format(16 + month)
			Sheet2.cell(row = 30, column = 45).value = "=sum(AS17:AS28)".format(16 + month)
			Sheet2.cell(row = 30, column = 46).value = "=sum(AT17:AT28)".format(16 + month)
			Sheet2.cell(row = 30, column = 47).value = "=sum(AU17:AU28)".format(16 + month)
			Sheet2.cell(row = 30, column = 48).value = "=sum(AV17:AV28)".format(16 + month)
			Sheet2.cell(row = 30, column = 49).value = "=sum(AW17:AW28)".format(16 + month)
			Sheet2.cell(row = 30, column = 50).value = "=sum(AX17:AX28)".format(16 + month)

			Sheet2.cell(row=34, column=21).value = "=SUMIF('Lead Schedule'!D:D,4427,'Lead Schedule'!I:I)".format(16 + month)
			Sheet2.cell(row=35, column=21).value = "=V30-U34".format(16 + month)

			Sheet2.cell(row=34, column=41).value = "=SUMIF('Lead Schedule'!D:D,4426,'Lead Schedule'!I:I)".format(16 + month)
			Sheet2.cell(row=35, column=41).value = "=AP30-AO34".format(16 + month)

			Sheet2.cell(row=37, column=41).value = "=SUMIF('Lead Schedule'!D:D,4423,'Lead Schedule'!I:I)".format(16 + month)
			Sheet2.cell(row=38, column=41).value = "=AS30+AO37".format(16 + month)

			Sheet2.cell(row=43, column=7).value = "=Q30-L30".format(16 + month)
			Sheet2.cell(row=44, column=7).value = "=SUM(G45:G49)".format(16 + month)
			Sheet2.cell(row=45, column=7).value = "=-SUMIF('TB Robot'!I:I,70,'TB Robot'!H:H)".format(16 + month)
			Sheet2.cell(row=46, column=7).value = "=SUMIF('TB Robot'!A:A,419,'TB Robot'!G:G)-SUMIF('TB Robot'!A:A,419,'TB Robot'!F:F)".format(16 + month)
			Sheet2.cell(row=47, column=7).value = "=SUMIF('TB Robot'!A:A,418,'TB Robot'!G:G)-SUMIF('TB Robot'!A:A,418,'TB Robot'!F:F)".format(16 + month)
			Sheet2.cell(row=48, column=7).value = "=SUMIF('TB Robot'!A:A,472,'TB Robot'!G:G)-SUMIF('TB Robot'!A:A,472,'TB Robot'!F:F)".format(16 + month)
			Sheet2.cell(row=49, column=7).value = "=SUMIF('TB Robot'!A:A,758,'TB Robot'!H:H)".format(16 + month)
			Sheet2.cell(row=50, column=7).value = "=SUM(G45:G49)".format(16 + month)
			Sheet2.cell(row=52, column=7).value = "=G43-G44".format(16 + month)
			Sheet2.cell(row=53, column=7).value = "=G52*19%".format(16 + month)

			Sheet2.cell(row=58, column=7).value = "=AL30".format(16 + month)
			Sheet2.cell(row=59, column=7).value = "=SUM(G60:G65)".format(16 + month)
			Sheet2.cell(row=60, column=7).value = "=-SUMIF('TB Robot'!I:I,60,'TB Robot'!H:H)".format(16 + month)
			Sheet2.cell(row=61, column=7).value = "=SUMIF('TB Robot'!J:J,3,'TB Robot'!H:H)-SUMIF('TB Robot'!I:I,39,'TB Robot'!H:H)".format(16 + month)
			Sheet2.cell(row=62, column=7).value = "0".format(16 + month)
			Sheet2.cell(row=63, column=7).value = "=SUMIF('TB Robot'!I:I,61,'TB Robot'!H:H)".format(16 + month)
			Sheet2.cell(row=64, column=7).value = "=SUMIF('TB Robot'!I:I,62,'TB Robot'!H:H)".format(16 + month)
			Sheet2.cell(row=65, column=7).value = "=SUMIF('TB Robot'!A:A,658,'TB Robot'!H:H)".format(16 + month)
			# Sheet2.cell(row=66, column=7).value = "Total"

			Sheet2.cell(row=68, column=7).value = "=G58-G59".format(16 + month)
			Sheet2.cell(row=69, column=7).value = "=G68*19%".format(16 + month)

		#design
		Sheet2.cell(row = 1, column = 2).font =ft1
		Sheet2.cell(row = 2, column = 2).font =ft1
		Sheet2.cell(row = 1, column = 10).font =ft1
		Sheet2.cell(row = 2, column = 10).font =ft1
		Sheet2.cell(row = 3, column = 10).font =ft1

		Sheet2.cell(row = 5, column = 2).font =f_testname
		Sheet2.cell(row = 7, column = 2).font =ft1
		Sheet2.cell(row = 11, column = 2).font =ft1
		Sheet2.cell(row = 12, column = 2).font =ft1

		for row in Sheet2['B17:BB30']:
			for cell in row:
				cell.font = font_worksheet

		for row in Sheet2['B16:C16']:
			for cell in row:
				cell.fill = cap_tabel_color_GT_movinchis

		Sheet2.cell(row=16, column=5).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=7).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=9).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=11).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=12).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=14).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=17).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=21).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=25).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=27).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=29).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=31).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=33).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=34).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=37).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=41).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=45).fill=cap_tabel_color_GT_movinchis
		Sheet2.cell(row=16, column=48).fill=cap_tabel_color_GT_movinchis

		Sheet2.cell(row=16, column=6).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=8).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=10).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=13).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=15).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=18).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=19).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=22).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=23).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=26).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=28).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=30).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=32).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=35).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=38).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=39).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=42).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=43).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=46).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=47).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=49).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=50).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=52).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=53).fill=cap_tabel_color_GT_movdeschis
		Sheet2.cell(row=16, column=54).fill=cap_tabel_color_GT_movdeschis

		#tables
		for row in Sheet2['A30:AW30']:
			for cell in row:
				cell.font = ft1

		for row in Sheet2['B16:BB16']:
			for cell in row:
				cell.font = cap_tabel

		for row in Sheet2['B17:AY30']:
			for cell in row:
				cell.number_format = '#,##0_);(#,##0)'

		for row in Sheet2['B16:C28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['E16:O28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['Q16:S28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['U16:W28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['Y16:AI28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['AK16:AM28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['AO16:AQ28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['AS16:AU28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['AV16:AX28']:
			for cell in row:
				cell.border = border

		for row in Sheet2['AZ16:BB28']:
			for cell in row:
				cell.border = border

		#COLORS
		for row in Sheet2['S17:S28']:
			for cell in row:
				cell.font = check_font
		Sheet2.cell(row=30,column=19).font = check_font

		for row in Sheet2['W17:W28']:
			for cell in row:
				cell.font = check_font
		Sheet2.cell(row=30,column=23).font = check_font

		for row in Sheet2['AM17:AM28']:
			for cell in row:
				cell.font = check_font
		Sheet2.cell(row=30,column=39).font = check_font

		for row in Sheet2['AQ17:AQ28']:
			for cell in row:
				cell.font = check_font
		Sheet2.cell(row=30,column=43).font = check_font

		for row in Sheet2['AX17:AX30']:
			for cell in row:
				cell.font = check_font
		Sheet2.cell(row=30,column=50).font = check_font

		for row in Sheet2['E43:G69']:
			for cell in row:
				cell.font = font_worksheet

		for row in Sheet2['G43:G69']:
			for cell in row:
				cell.number_format = '#,##0_);(#,##0)'

		for row in Sheet2['E43:G44']:
			for cell in row:
				cell.font = ft1

		for row in Sheet2['F50:G50']:
			for cell in row:
				cell.font = ft1

		for row in Sheet2['E52:G52']:
			for cell in row:
				cell.font = check_font

		for row in Sheet2['F58:G59']:
			for cell in row:
				cell.font = ft1

		for row in Sheet2['F66:G66']:
			for cell in row:
				cell.font = ft1

		for row in Sheet2['F68:G68']:
			for cell in row:
				cell.font = check_font

		Sheet2.cell(row=34, column=21).number_format = '#,##0_);(#,##0)'
		Sheet2.cell(row=35, column=21).number_format = '#,##0_);(#,##0)'

		Sheet2.cell(row=34, column=41).number_format = '#,##0_);(#,##0)'
		Sheet2.cell(row=35, column=41).number_format = '#,##0_);(#,##0)'

		Sheet2.cell(row=37, column=41).number_format = '#,##0_);(#,##0)'
		Sheet2.cell(row=38, column=41).number_format = '#,##0_);(#,##0)'

		#column dimensions
		Sheet2.column_dimensions['C'].width = 20
		Sheet2.column_dimensions['E'].width = 20
		Sheet2.column_dimensions['F'].width = 20
		Sheet2.column_dimensions['G'].width = 20
		Sheet2.column_dimensions['H'].width = 20
		Sheet2.column_dimensions['I'].width = 20
		Sheet2.column_dimensions['J'].width = 20
		Sheet2.column_dimensions['K'].width = 20
		Sheet2.column_dimensions['L'].width = 20
		Sheet2.column_dimensions['M'].width = 20
		Sheet2.column_dimensions['N'].width = 20
		Sheet2.column_dimensions['O'].width = 20
		Sheet2.column_dimensions['Q'].width = 20
		Sheet2.column_dimensions['R'].width = 20
		Sheet2.column_dimensions['S'].width = 20
		Sheet2.column_dimensions['U'].width = 20
		Sheet2.column_dimensions['V'].width = 20
		Sheet2.column_dimensions['W'].width = 20
		Sheet2.column_dimensions['Y'].width = 20
		Sheet2.column_dimensions['Z'].width = 20
		Sheet2.column_dimensions['AA'].width = 20
		Sheet2.column_dimensions['AB'].width = 20
		Sheet2.column_dimensions['AC'].width = 20
		Sheet2.column_dimensions['AD'].width = 20
		Sheet2.column_dimensions['AE'].width = 20
		Sheet2.column_dimensions['AF'].width = 20
		Sheet2.column_dimensions['AG'].width = 20
		Sheet2.column_dimensions['AH'].width = 20
		Sheet2.column_dimensions['AI'].width = 20
		Sheet2.column_dimensions['AK'].width = 20
		Sheet2.column_dimensions['AL'].width = 20
		Sheet2.column_dimensions['AM'].width = 20
		Sheet2.column_dimensions['AO'].width = 20
		Sheet2.column_dimensions['AP'].width = 20
		Sheet2.column_dimensions['AQ'].width = 20
		Sheet2.column_dimensions['AS'].width = 20
		Sheet2.column_dimensions['AT'].width = 20
		Sheet2.column_dimensions['AU'].width = 20
		Sheet2.column_dimensions['AV'].width = 20
		Sheet2.column_dimensions['AW'].width = 30
		Sheet2.column_dimensions['AX'].width = 20
		Sheet2.column_dimensions['AY'].width = 20
		Sheet2.column_dimensions['AZ'].width = 20
		Sheet2.column_dimensions['BA'].width = 20
		Sheet2.column_dimensions['BB'].width = 20

		Sheet2['B16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['C16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['E16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['F16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['G16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['H16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['I16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['J16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['K16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['L16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['M16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['N16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['O16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['Q16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['R16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['S16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['U16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['V16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['W16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['Y16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['Z16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AA16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AB16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AC16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AD16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AE16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AF16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AG16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AH16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AI16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AK16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AL16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AM16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AO16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AP16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AQ16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AS16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AT16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AU16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AV16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AW16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AX16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AY16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['AZ16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['BA16'].alignment = Alignment(wrapText=True, vertical='center')
		Sheet2['BB16'].alignment = Alignment(wrapText=True, vertical='center')

		Sheet3 = wb.create_sheet("VAT Test 2")
		Sheet3.sheet_view.showGridLines = False

		Sheet3.cell(row = 1, column = 1).value ="Client Name:"
		Sheet3.cell(row = 2, column = 1).value ="Period End:"
		Sheet3.cell(row = 1, column = 14).value ="Prepared By:" 
		Sheet3.cell(row = 2, column = 14).value ="Date:"
		Sheet3.cell(row = 3, column = 14).value ="Reviewed by:"

		Sheet3.cell(row = 4, column = 1).value ="VAT Reconciliation Summary"
		Sheet3.cell(row = 6, column = 1).value ="Objective:"
		Sheet3.cell(row = 8, column = 1).value ="Work Done:"
		Sheet3.cell(row = 13, column = 1).value ="Findings:"
		Sheet3.cell(row = 15, column = 1).value ="Conclusion:"

		Sheet3.cell(row = 18, column = 1).value ="As per VAT Statement (300)"
		Sheet3.cell(row = 27, column = 1).value ="As per Monthly TB"
		Sheet3.cell(row = 32, column = 1).value ="Check Statement 300 vs Monthly TB"

		Sheet3.cell(row = 37, column = 1).value ="As per Sales and Acquisitions Journals"
		Sheet3.cell(row = 43, column = 1).value ="Check Journals vs Trial Balance"
		Sheet3.cell(row = 49, column = 1).value ="Check Statement 300 vs Journals"

		Sheet3.cell(row = 6, column = 2).value ="To test the completeness of the VAT balances."
		Sheet3.cell(row = 8, column = 2).value ="We have obtained the Monthly VAT returns prepared for Local Budget (no. 300)."
		Sheet3.cell(row = 9, column = 2).value ="We have obtained the monthly sales and acquisition journals in order to extract the VAT and reconciled with the value presented in VAT statement. "
		Sheet3.cell(row = 10, column = 2).value ="We have reconciled the VAT from the VAT statement and VAT from the sales and acquisition journals with the value from trial balance."

		#bilding tables
		#table 1
		Sheet3.cell(row = 18, column = 2).value ="Type of VAT"
		Sheet3.cell(row = 18, column = 3).value ="January"
		Sheet3.cell(row = 18, column = 4).value ="February"
		Sheet3.cell(row = 18, column = 5).value ="March"
		Sheet3.cell(row = 18, column = 6).value ="April"
		Sheet3.cell(row = 18, column = 7).value ="May"
		Sheet3.cell(row = 18, column = 8).value ="June"
		Sheet3.cell(row = 18, column = 9).value ="July"
		Sheet3.cell(row = 18, column = 10).value ="August"
		Sheet3.cell(row = 18, column = 11).value ="September"
		Sheet3.cell(row = 18, column = 12).value ="October"
		Sheet3.cell(row = 18, column = 13).value ="November"
		Sheet3.cell(row = 18, column = 14).value ="December"
		Sheet3.cell(row = 18, column = 15).value ="Total"

		Sheet3.cell(row = 19, column = 2).value ="VAT Collectable"
		Sheet3.cell(row = 20, column = 2).value ="VAT Deductible"
		# Sheet3.cell(row = 21, column = 2).value ="Reverse charge"
		# Sheet3.cell(row = 22, column = 2).value ="Exempted"
		# Sheet3.cell(row = 23, column = 2).value ="Regularization collectable"
		Sheet3.cell(row = 24, column = 2).value ="Regularization deduction"
		Sheet3.cell(row = 25, column = 2).value ="VAT Payable/ recoverable"

		#table2
		Sheet3.cell(row = 27, column = 2).value ="Type of VAT"
		Sheet3.cell(row = 27, column = 3).value ="January"
		Sheet3.cell(row = 27, column = 4).value ="February"
		Sheet3.cell(row = 27, column = 5).value ="March"
		Sheet3.cell(row = 27, column = 6).value ="April"
		Sheet3.cell(row = 27, column = 7).value ="May"
		Sheet3.cell(row = 27, column = 8).value ="June"
		Sheet3.cell(row = 27, column = 9).value ="July"
		Sheet3.cell(row = 27, column = 10).value ="August"
		Sheet3.cell(row = 27, column = 11).value ="September"
		Sheet3.cell(row = 27, column = 12).value ="October"
		Sheet3.cell(row = 27, column = 13).value ="November"
		Sheet3.cell(row = 27, column = 14).value ="December"
		Sheet3.cell(row = 27, column = 15).value ="Total"

		Sheet3.cell(row = 28, column = 2).value ="VAT Collectable 4427"
		Sheet3.cell(row = 29, column = 2).value ="VAT Deductible 4426"
		Sheet3.cell(row = 30, column = 2).value ="VAT Payable/ recoverable"

		#table3
		Sheet3.cell(row = 32, column = 2).value ="Type of VAT"
		Sheet3.cell(row = 32, column = 3).value ="January"
		Sheet3.cell(row = 32, column = 4).value ="February"
		Sheet3.cell(row = 32, column = 5).value ="March"
		Sheet3.cell(row = 32, column = 6).value ="April"
		Sheet3.cell(row = 32, column = 7).value ="May"
		Sheet3.cell(row = 32, column = 8).value ="June"
		Sheet3.cell(row = 32, column = 9).value ="July"
		Sheet3.cell(row = 32, column = 10).value ="August"
		Sheet3.cell(row = 32, column = 11).value ="September"
		Sheet3.cell(row = 32, column = 12).value ="October"
		Sheet3.cell(row = 32, column = 13).value ="November"
		Sheet3.cell(row = 32, column = 14).value ="December"
		Sheet3.cell(row = 32, column = 15).value ="Total"

		Sheet3.cell(row = 33, column = 2).value ="VAT Collectable 4427"
		Sheet3.cell(row = 34, column = 2).value ="VAT Deductible 4426"
		Sheet3.cell(row = 35, column = 2).value ="VAT Payable/ recoverable"

		#table4
		Sheet3.cell(row = 37, column = 2).value ="Type of VAT"
		Sheet3.cell(row = 37, column = 3).value ="January"
		Sheet3.cell(row = 37, column = 4).value ="February"
		Sheet3.cell(row = 37, column = 5).value ="March"
		Sheet3.cell(row = 37, column = 6).value ="April"
		Sheet3.cell(row = 37, column = 7).value ="May"
		Sheet3.cell(row = 37, column = 8).value ="June"
		Sheet3.cell(row = 37, column = 9).value ="July"
		Sheet3.cell(row = 37, column = 10).value ="August"
		Sheet3.cell(row = 37, column = 11).value ="September"
		Sheet3.cell(row = 37, column = 12).value ="October"
		Sheet3.cell(row = 37, column = 13).value ="November"
		Sheet3.cell(row = 37, column = 14).value ="December"
		Sheet3.cell(row = 37, column = 15).value ="Total"

		Sheet3.cell(row = 38, column = 2).value ="VAT Collectable 4427"
		Sheet3.cell(row = 39, column = 2).value ="VAT Deductible 4426"
		Sheet3.cell(row = 40, column = 2).value ="VAT Payable/ recoverable"

		#table5
		Sheet3.cell(row = 43, column = 2).value ="Type of VAT"
		Sheet3.cell(row = 43, column = 3).value ="January"
		Sheet3.cell(row = 43, column = 4).value ="February"
		Sheet3.cell(row = 43, column = 5).value ="March"
		Sheet3.cell(row = 43, column = 6).value ="April"
		Sheet3.cell(row = 43, column = 7).value ="May"
		Sheet3.cell(row = 43, column = 8).value ="June"
		Sheet3.cell(row = 43, column = 9).value ="July"
		Sheet3.cell(row = 43, column = 10).value ="August"
		Sheet3.cell(row = 43, column = 11).value ="September"
		Sheet3.cell(row = 43, column = 12).value ="October"
		Sheet3.cell(row = 43, column = 13).value ="November"
		Sheet3.cell(row = 43, column = 14).value ="December"
		Sheet3.cell(row = 43, column = 15).value ="Total"

		Sheet3.cell(row = 44, column = 2).value ="VAT Collectable 4427"
		Sheet3.cell(row = 45, column = 2).value ="VAT Deductible 4426"
		Sheet3.cell(row = 46, column = 2).value ="VAT Payable/ recoverable"

		#table6
		Sheet3.cell(row = 49, column = 2).value ="Type of VAT"
		Sheet3.cell(row = 49, column = 3).value ="January"
		Sheet3.cell(row = 49, column = 4).value ="February"
		Sheet3.cell(row = 49, column = 5).value ="March"
		Sheet3.cell(row = 49, column = 6).value ="April"
		Sheet3.cell(row = 49, column = 7).value ="May"
		Sheet3.cell(row = 49, column = 8).value ="June"
		Sheet3.cell(row = 49, column = 9).value ="July"
		Sheet3.cell(row = 49, column = 10).value ="August"
		Sheet3.cell(row = 49, column = 11).value ="September"
		Sheet3.cell(row = 49, column = 12).value ="October"
		Sheet3.cell(row = 49, column = 13).value ="November"
		Sheet3.cell(row = 49, column = 14).value ="December"
		Sheet3.cell(row = 49, column = 15).value ="Total"

		Sheet3.cell(row = 50, column = 2).value ="VAT Collectable 4427"
		Sheet3.cell(row = 51, column = 2).value ="VAT Deductible 4426"
		Sheet3.cell(row = 52, column = 2).value ="VAT Payable/ recoverable"

		#content
		Sheet3.cell(row = 1, column = 2).value =clientname1
		Sheet3.cell(row = 2, column = 2).value =yearEnd1
		Sheet3.cell(row = 2, column = 2).number_format='mm/dd/yyyy'
		Sheet3.cell(row = 1, column = 15).value =preparedBy1
		Sheet3.cell(row = 2, column = 15).value =datePrepared1
		Sheet3.cell(row = 2, column = 15).number_format='mm/dd/yyyy'
		
		Sheet3.cell(row = 19, column = 3).value ="='VAT Test'!V17"
		Sheet3.cell(row = 19, column = 4).value ="='VAT Test'!V18"
		Sheet3.cell(row = 19, column = 5).value ="='VAT Test'!V19"
		Sheet3.cell(row = 19, column = 6).value ="='VAT Test'!V20"
		Sheet3.cell(row = 19, column = 7).value ="='VAT Test'!V21"
		Sheet3.cell(row = 19, column = 8).value ="='VAT Test'!V22"
		Sheet3.cell(row = 19, column = 9).value ="='VAT Test'!V23"
		Sheet3.cell(row = 19, column = 10).value ="='VAT Test'!V24"
		Sheet3.cell(row = 19, column = 11).value ="='VAT Test'!V25"
		Sheet3.cell(row = 19, column = 12).value ="='VAT Test'!V26"
		Sheet3.cell(row = 19, column = 13).value ="='VAT Test'!V27"
		Sheet3.cell(row = 19, column = 14).value ="='VAT Test'!V28"
		Sheet3.cell(row = 19, column = 15).value ="=SUM(C19:N19)"

		Sheet3.cell(row = 20, column = 3).value ="='VAT Test'!AP17"
		Sheet3.cell(row = 20, column = 4).value ="='VAT Test'!AP18"
		Sheet3.cell(row = 20, column = 5).value ="='VAT Test'!AP19"
		Sheet3.cell(row = 20, column = 6).value ="='VAT Test'!AP20"
		Sheet3.cell(row = 20, column = 7).value ="='VAT Test'!AP21"
		Sheet3.cell(row = 20, column = 8).value ="='VAT Test'!AP22"
		Sheet3.cell(row = 20, column = 9).value ="='VAT Test'!AP23"
		Sheet3.cell(row = 20, column = 10).value ="='VAT Test'!AP24"
		Sheet3.cell(row = 20, column = 11).value ="='VAT Test'!AP25"
		Sheet3.cell(row = 20, column = 12).value ="='VAT Test'!AP26"
		Sheet3.cell(row = 20, column = 13).value ="='VAT Test'!AP27"
		Sheet3.cell(row = 20, column = 14).value ="='VAT Test'!AP28"
		Sheet3.cell(row = 20, column = 15).value ="=SUM(C20:N20)"

		# Sheet3.cell(row = 22, column = 3).value ="='VAT Test'!K17"
		# Sheet3.cell(row = 22, column = 4).value ="='VAT Test'!K18"
		# Sheet3.cell(row = 22, column = 5).value ="='VAT Test'!K19"
		# Sheet3.cell(row = 22, column = 6).value ="='VAT Test'!K20"
		# Sheet3.cell(row = 22, column = 7).value ="='VAT Test'!K21"
		# Sheet3.cell(row = 22, column = 8).value ="='VAT Test'!K22"
		# Sheet3.cell(row = 22, column = 9).value ="='VAT Test'!K23"
		# Sheet3.cell(row = 22, column = 10).value ="='VAT Test'!K24"
		# Sheet3.cell(row = 22, column = 11).value ="='VAT Test'!K25"
		# Sheet3.cell(row = 22, column = 12).value ="='VAT Test'!K26"
		# Sheet3.cell(row = 22, column = 13).value ="='VAT Test'!K27"
		# Sheet3.cell(row = 22, column = 14).value ="='VAT Test'!K28"
		# Sheet3.cell(row = 22, column = 15).value ="=SUM(C22:N22)"

		Sheet3.cell(row = 25, column = 3).value ="=C24+C20-C19"
		Sheet3.cell(row = 25, column = 4).value ="=D24+D20-D19"
		Sheet3.cell(row = 25, column = 5).value ="=E24+E20-E19"
		Sheet3.cell(row = 25, column = 6).value ="=F24+F20-F19"
		Sheet3.cell(row = 25, column = 7).value ="=G24+G20-G19"
		Sheet3.cell(row = 25, column = 8).value ="=H24+H20-H19"
		Sheet3.cell(row = 25, column = 9).value ="=I24+I20-I19"
		Sheet3.cell(row = 25, column = 10).value ="=J24+J20-J19"
		Sheet3.cell(row = 25, column = 11).value ="=K24+K20-K19"
		Sheet3.cell(row = 25, column = 12).value ="=L24+L20-L19"
		Sheet3.cell(row = 25, column = 13).value ="=M24+M20-M19"
		Sheet3.cell(row = 25, column = 14).value ="=N24+N20-N19"
		Sheet3.cell(row = 25, column = 15).value ="=SUM(C25:N25)"

		#table3
		Sheet3.cell(row = 33, column = 3).value ="=C19-C28"
		Sheet3.cell(row = 33, column = 4).value ="=D19-D28"
		Sheet3.cell(row = 33, column = 5).value ="=E19-E28"
		Sheet3.cell(row = 33, column = 6).value ="=F19-F28"
		Sheet3.cell(row = 33, column = 7).value ="=G19-G28"
		Sheet3.cell(row = 33, column = 8).value ="=H19-H28"
		Sheet3.cell(row = 33, column = 9).value ="=I19-I28"
		Sheet3.cell(row = 33, column = 10).value ="=J19-J28"
		Sheet3.cell(row = 33, column = 11).value ="=K19-K28"
		Sheet3.cell(row = 33, column = 12).value ="=L19-L28"
		Sheet3.cell(row = 33, column = 13).value ="=M19-M28"
		Sheet3.cell(row = 33, column = 14).value ="=N19-N28"
		Sheet3.cell(row = 33, column = 15).value ="=SUM(C33:N33)"

		Sheet3.cell(row = 34, column = 3).value ="=C20+C24-C29"
		Sheet3.cell(row = 34, column = 4).value ="=D20+D24-D29"
		Sheet3.cell(row = 34, column = 5).value ="=E20+E24-E29"
		Sheet3.cell(row = 34, column = 6).value ="=F20+F24-F29"
		Sheet3.cell(row = 34, column = 7).value ="=G20+G24-G29"
		Sheet3.cell(row = 34, column = 8).value ="=H20+H24-H29"
		Sheet3.cell(row = 34, column = 9).value ="=I20+I24-I29"
		Sheet3.cell(row = 34, column = 10).value ="=J20+J24-J29"
		Sheet3.cell(row = 34, column = 11).value ="=K20+K24-K29"
		Sheet3.cell(row = 34, column = 12).value ="=L20+L24-L29"
		Sheet3.cell(row = 34, column = 13).value ="=M20+M24-M29"
		Sheet3.cell(row = 34, column = 14).value ="=N20+N24-N29"
		Sheet3.cell(row = 34, column = 15).value ="=SUM(C34:N34)"

		Sheet3.cell(row = 35, column = 3).value ="=SUM(C33:C34)"
		Sheet3.cell(row = 35, column = 4).value ="=SUM(D33:D34)"
		Sheet3.cell(row = 35, column = 5).value ="=SUM(E33:E34)"
		Sheet3.cell(row = 35, column = 6).value ="=SUM(F33:F34)"
		Sheet3.cell(row = 35, column = 7).value ="=SUM(G33:G34)"
		Sheet3.cell(row = 35, column = 8).value ="=SUM(H33:H34)"
		Sheet3.cell(row = 35, column = 9).value ="=SUM(I33:I34)"
		Sheet3.cell(row = 35, column = 10).value ="=SUM(J33:J34)"
		Sheet3.cell(row = 35, column = 11).value ="=SUM(K33:K34)"
		Sheet3.cell(row = 35, column = 12).value ="=SUM(L33:L34)"
		Sheet3.cell(row = 35, column = 13).value ="=SUM(M33:M34)"
		Sheet3.cell(row = 35, column = 14).value ="=SUM(N33:N34)"
		Sheet3.cell(row = 35, column = 15).value ="=SUM(C35:N35)"

		#table 5
		Sheet3.cell(row = 44, column = 3).value ="=C38-C28"
		Sheet3.cell(row = 44, column = 4).value ="=D38-D28"
		Sheet3.cell(row = 44, column = 5).value ="=E38-E28"
		Sheet3.cell(row = 44, column = 6).value ="=F38-F28"
		Sheet3.cell(row = 44, column = 7).value ="=G38-G28"
		Sheet3.cell(row = 44, column = 8).value ="=H38-H28"
		Sheet3.cell(row = 44, column = 9).value ="=I38-I28"
		Sheet3.cell(row = 44, column = 10).value ="=J38-J28"
		Sheet3.cell(row = 44, column = 11).value ="=K38-K28"
		Sheet3.cell(row = 44, column = 12).value ="=L38-L28"
		Sheet3.cell(row = 44, column = 13).value ="=M38-M28"
		Sheet3.cell(row = 44, column = 14).value ="=N38-N28"
		Sheet3.cell(row = 44, column = 15).value ="=SUM(C44:N44)"

		Sheet3.cell(row = 45, column = 3).value ="=C39-C29"
		Sheet3.cell(row = 45, column = 4).value ="=D39-D29"
		Sheet3.cell(row = 45, column = 5).value ="=E39-E29"
		Sheet3.cell(row = 45, column = 6).value ="=F39-F29"
		Sheet3.cell(row = 45, column = 7).value ="=G39-G29"
		Sheet3.cell(row = 45, column = 8).value ="=H39-H29"
		Sheet3.cell(row = 45, column = 9).value ="=I39-I29"
		Sheet3.cell(row = 45, column = 10).value ="=J39-J29"
		Sheet3.cell(row = 45, column = 11).value ="=K39-K29"
		Sheet3.cell(row = 45, column = 12).value ="=L39-L29"
		Sheet3.cell(row = 45, column = 13).value ="=M39-M29"
		Sheet3.cell(row = 45, column = 14).value ="=N39-N29"
		Sheet3.cell(row = 45, column = 15).value ="=SUM(C45:N45)"

		Sheet3.cell(row = 46, column = 3).value ="=C40-C30"
		Sheet3.cell(row = 46, column = 4).value ="=D40-D30"
		Sheet3.cell(row = 46, column = 5).value ="=E40-E30"
		Sheet3.cell(row = 46, column = 6).value ="=F40-F30"
		Sheet3.cell(row = 46, column = 7).value ="=G40-G30"
		Sheet3.cell(row = 46, column = 8).value ="=H40-H30"
		Sheet3.cell(row = 46, column = 9).value ="=I40-I30"
		Sheet3.cell(row = 46, column = 10).value ="=J40-J30"
		Sheet3.cell(row = 46, column = 11).value ="=K40-K30"
		Sheet3.cell(row = 46, column = 12).value ="=L40-L30"
		Sheet3.cell(row = 46, column = 13).value ="=M40-M30"
		Sheet3.cell(row = 46, column = 14).value ="=N40-N30"
		Sheet3.cell(row = 46, column = 15).value ="=SUM(C46:N46)"

		#table6
		Sheet3.cell(row = 50, column = 3).value ="=C19-C38"
		Sheet3.cell(row = 50, column = 4).value ="=D19-D38"
		Sheet3.cell(row = 50, column = 5).value ="=E19-E38"
		Sheet3.cell(row = 50, column = 6).value ="=F19-F38"
		Sheet3.cell(row = 50, column = 7).value ="=G19-G38"
		Sheet3.cell(row = 50, column = 8).value ="=H19-H38"
		Sheet3.cell(row = 50, column = 9).value ="=I19-I38"
		Sheet3.cell(row = 50, column = 10).value ="=J19-J38"
		Sheet3.cell(row = 50, column = 11).value ="=K19-K38"
		Sheet3.cell(row = 50, column = 12).value ="=L19-L38"
		Sheet3.cell(row = 50, column = 13).value ="=M19-M38"
		Sheet3.cell(row = 50, column = 14).value ="=N19-N38"
		Sheet3.cell(row = 50, column = 15).value ="=SUM(C50:N50)"

		Sheet3.cell(row = 51, column = 3).value ="=C20+C24-C39"
		Sheet3.cell(row = 51, column = 4).value ="=D20+D24-D39"
		Sheet3.cell(row = 51, column = 5).value ="=E20+E24-E39"
		Sheet3.cell(row = 51, column = 6).value ="=F20+F24-F39"
		Sheet3.cell(row = 51, column = 7).value ="=G20+G24-G39"
		Sheet3.cell(row = 51, column = 8).value ="=H20+H24-H39"
		Sheet3.cell(row = 51, column = 9).value ="=I20+I24-I39"
		Sheet3.cell(row = 51, column = 10).value ="=J20+J24-J39"
		Sheet3.cell(row = 51, column = 11).value ="=K20+K24-K39"
		Sheet3.cell(row = 51, column = 12).value ="=L20+L24-L39"
		Sheet3.cell(row = 51, column = 13).value ="=M20+M24-M39"
		Sheet3.cell(row = 51, column = 14).value ="=N20+N24-N39"
		Sheet3.cell(row = 51, column = 15).value ="=SUM(C51:N51)"

		Sheet3.cell(row = 52, column = 3).value ="=C25-C40"
		Sheet3.cell(row = 52, column = 4).value ="=D25-D40"
		Sheet3.cell(row = 52, column = 5).value ="=E25-E40"
		Sheet3.cell(row = 52, column = 6).value ="=F25-F40"
		Sheet3.cell(row = 52, column = 7).value ="=G25-G40"
		Sheet3.cell(row = 52, column = 8).value ="=H25-H40"
		Sheet3.cell(row = 52, column = 9).value ="=I25-I40"
		Sheet3.cell(row = 52, column = 10).value ="=J25-J40"
		Sheet3.cell(row = 52, column = 11).value ="=K25-K40"
		Sheet3.cell(row = 52, column = 12).value ="=L25-L40"
		Sheet3.cell(row = 52, column = 13).value ="=M25-M40"
		Sheet3.cell(row = 52, column = 14).value ="=N25-N40"
		Sheet3.cell(row = 52, column = 15).value ="=SUM(C52:N52)"

		#format
		for row in Sheet3['C19:O52']:
			for cell in row:
				cell.number_format = '#,##0_);(#,##0)'

		for row in Sheet3['B18:O25']:
			for cell in row:
				cell.border = border

		for row in Sheet3['B27:O30']:
			for cell in row:
				cell.border = border

		for row in Sheet3['B32:O35']:
			for cell in row:
				cell.border = border

		for row in Sheet3['B37:O40']:
			for cell in row:
				cell.border = border

		for row in Sheet3['B43:O46']:
			for cell in row:
				cell.border = border

		for row in Sheet3['B49:O52']:
			for cell in row:
				cell.border = border

		#COLORS
		Sheet3.cell(row = 1, column = 1).font =ft1
		Sheet3.cell(row = 2, column = 1).font =ft1
		Sheet3.cell(row = 1, column = 14).font =ft1
		Sheet3.cell(row = 2, column = 14).font =ft1
		Sheet3.cell(row = 3, column = 14).font =ft1

		Sheet3.cell(row = 4, column = 1).font = f_testname
		Sheet3.cell(row = 6, column = 1).font =ft1
		Sheet3.cell(row = 8, column = 1).font =ft1
		Sheet3.cell(row = 13, column = 1).font =ft1
		Sheet3.cell(row = 15, column = 1).font =ft1

		Sheet3.cell(row = 18, column = 1).font =ft1
		Sheet3.cell(row = 27, column = 1).font =blue_bold_font
		Sheet3.cell(row = 32, column = 1).font =check_font

		Sheet3.cell(row = 37, column = 1).font =ft1
		Sheet3.cell(row = 43, column = 1).font =check_font
		Sheet3.cell(row = 49, column = 1).font =check_font

		#table2 content
		for row in Sheet3['B28:O29']:
			for cell in row:
				cell.font = blue_thin_font

		#table3 content
		for row in Sheet3['C33:O35']:
			for cell in row:
				cell.font = check_font_1

		#table5 content
		for row in Sheet3['C44:O46']:
			for cell in row:
				cell.font = check_font_1

		#table6 content
		for row in Sheet3['C50:O52']:
			for cell in row:
				cell.font = check_font_1

		#header months
		for row in Sheet3['B18:O18']:
			for cell in row:
				cell.font = ft1

		for row in Sheet3['B25:O25']:
			for cell in row:
				cell.font = ft1

		for row in Sheet3['B27:O27']:
			for cell in row:
				cell.font = blue_bold_font

		for row in Sheet3['B32:O32']:
			for cell in row:
				cell.font = ft1

		for row in Sheet3['B37:O37']:
			for cell in row:
				cell.font = ft1

		for row in Sheet3['B43:O43']:
			for cell in row:
				cell.font = ft1

		for row in Sheet3['B49:O49']:
			for cell in row:
				cell.font = ft1

		#totals
		for row in Sheet3['B25:O25']:
			for cell in row:
				cell.font = ft1

		for row in Sheet3['B30:O30']:
			for cell in row:
				cell.font = blue_bold_font
		Sheet3.cell(row=35,column=2).font=ft1

		for row in Sheet3['B40:O40']:
			for cell in row:
				cell.font = ft1
		Sheet3.cell(row=46,column=2).font=ft1

		Sheet3.column_dimensions['A'].width = 38
		Sheet3.column_dimensions['B'].width = 26
		Sheet3.column_dimensions['C'].width = 13
		Sheet3.column_dimensions['D'].width = 13
		Sheet3.column_dimensions['E'].width = 13
		Sheet3.column_dimensions['F'].width = 13
		Sheet3.column_dimensions['G'].width = 13
		Sheet3.column_dimensions['H'].width = 13
		Sheet3.column_dimensions['I'].width = 13
		Sheet3.column_dimensions['J'].width = 13
		Sheet3.column_dimensions['K'].width = 13
		Sheet3.column_dimensions['L'].width = 13
		Sheet3.column_dimensions['M'].width = 13
		Sheet3.column_dimensions['N'].width = 13
		Sheet3.column_dimensions['O'].width = 13
		Sheet3.column_dimensions['Q'].width = 13

		Sheet4 = wb.create_sheet("TB Robot")
		Sheet4.sheet_view.showGridLines = False

		tb = openpyxl.load_workbook(file_TB, data_only = 'True') #deschidem TB-ul
		tb1 = tb.active

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "Account":
					rand_tb = cell.row
					coloana_acc_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			account = [b.value for b in tb1[coloana_acc_tb][rand_tb:lun]]           
		except:
			flash("Please insert the correct header for Account in Trial Balance file")
			return render_template("index.html")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "Description":
					rand_tb = cell.row
					coloana_descr_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			descr = [b.value for b in tb1[coloana_descr_tb][rand_tb:lun]]          
		except:
			flash("Please insert the correct header for Description in Trial Balance file")
			return render_template("index.html")


		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "OB":
					rand_tb = cell.row
					coloana_opTB_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			opTB = [b.value for b in tb1[coloana_opTB_tb][rand_tb:lun]]          
		except:
			flash("Please insert the correct header for OB in Trial Balance file")
			return render_template("index.html")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "CM":
					rand_tb = cell.row
					coloana_cr_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			cr_mv = [b.value for b in tb1[coloana_cr_tb][rand_tb:lun]]         
		except:
			flash("Please insert the correct header for CM in Trial Balance file")
			return render_template("index.html")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "DM":
					rand_tb = cell.row
					coloana_db_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			dr_mv = [b.value for b in tb1[coloana_db_tb][rand_tb:lun]]         
		except:
			flash("Please insert the correct header for DM in Trial Balance file")
			return render_template("index.html")

		for row in tb1.iter_rows():
			for cell in row:
				if cell.value == "CB":
					rand_tb = cell.row
					coloana_clTB_tb = cell.column
					lun = len(tb1[cell.column])
		try:
			clTB = [b.value for b in tb1[coloana_clTB_tb][rand_tb:lun]]       
		except:
			flash("Please insert the correct header for CB in Trial Balance file")
			return render_template("index.html")

		Sheet4.cell(row = 1, column = 1).value = "Synt(3)"
		Sheet4.cell(row = 1, column = 2).value = "Synt(4)"
		Sheet4.cell(row = 1, column = 3).value = "Account"
		Sheet4.cell(row = 1, column = 4).value = "Description"
		Sheet4.cell(row = 1, column = 5).value = "Opening Balance"
		Sheet4.cell(row = 1, column = 6).value = "Debit Movement"
		Sheet4.cell(row = 1, column = 7).value = "Credit Movement"
		Sheet4.cell(row = 1, column = 8).value = "Closing Balance"
		Sheet4.cell(row = 1, column = 9).value = "Synt(2)"
		Sheet4.cell(row=1, column=10).value = "Class"

		# # ....adaugi tu restul adica: synt 4, account,descript,ob,dm,cm,cbp
		for i in range(0,len(account)):
			Sheet4.cell(row = 2 + i, column = 3).value = account[i]
			Sheet4.cell(row = 2 + i, column = 2).value = str(account[i])[:4]   #in Excel =left("celula", 4)

		for i in range(0, len(account)):
			Sheet4.cell(row = 2 + i, column = 1).value =  str(account[i])[:3] #in Excel =left("celula", 3)

		for i in range(0, len(descr)):
			Sheet4.cell(row = 2 + i, column = 4).value = descr[i]

		for i in range(0, len(opTB)):
			Sheet4.cell(row = 2 + i, column = 5).value = opTB[i]

		for i in range(0, len(cr_mv)):
			Sheet4.cell(row = 2 + i, column = 6).value = cr_mv[i]

		for i in range(0, len(dr_mv)):
			Sheet4.cell(row = 2 + i, column = 7).value = dr_mv[i]

		for i in range(0, len(clTB)):
			Sheet4.cell(row = 2 + i, column = 8).value = clTB[i]

		for i in range(0, len(account)):
			Sheet4.cell(row = 2 + i, column = 9).value =  str(account[i])[:2]

		for i in range(0, len(account)):
			Sheet4.cell(row=2 + i, column=10).value = str(account[i])[:1]

		Sheet5 = wb.create_sheet("TB PBC")

		mr = tb1.max_row
		mc = tb1.max_column
		# copying the cell values from source
		# excel file to destination excel file
		for i in range (1, mr + 1):
			for j in range (1, mc + 1):
		# reading cell value from source excel file
				c = tb1.cell(row = i, column = j)
		# writing the read value to destination excel file
				Sheet5.cell(row = i, column = j).value = c.value

		std = wb["Sheet"]
		wb.remove(std)
		folderpath = "/home/fsbot/storage/vat"
		file_pathFS = os.path.join(folderpath, "VAT test"+" "+clientname1+" "+str(denis)+".xlsx")
		wb.save(file_pathFS)
		# out.save(folderpath + "/" + "T10 - VAT Test" + ".xlsx")
		return send_from_directory(folderpath, "VAT Test" + " " + clientname1 +" "+str(denis)+ ".xlsx", as_attachment=True)

	return render_template("VAT.html")
		
if __name__ == '__main__':
   	app.run()
